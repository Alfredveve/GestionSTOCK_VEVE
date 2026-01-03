import sys
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(headers, data, title, filename_prefix="Export"):
    """
    Standardized utility for premium Excel exports.
    """
    # Hack: Bypass broken lxml on some Windows/Python 3.12 environments
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = title[:30] # Excel sheet title limit

        # Colors (Metronic style)
        PRIMARY_BLUE = "009EF7"
        WHITE = "FFFFFF"
        LIGHT_GRAY = "F5F8FA"
        BORDER_COLOR = "E4E6EF"

        # Styles
        header_fill = PatternFill(start_color=PRIMARY_BLUE, end_color=PRIMARY_BLUE, fill_type="solid")
        header_font = Font(bold=True, color=WHITE, size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )

        # Write Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Detection of currency columns
        currency_keywords = ['prix', 'total', 'montant', 'solde', 'valeur', 'bénéfice', 'intérêt', 'coût', 'remise', 'dépense', 'marge']
        currency_cols = []
        for col_num, header in enumerate(headers, 1):
            if any(keyword in header.lower() for keyword in currency_keywords):
                currency_cols.append(col_num)

        # Write Data
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = thin_border
                
                if col_num in currency_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # GNF format with thousands separator
                    cell.number_format = '#,##0 "GNF"'
                else:
                    cell.alignment = left_alignment
                
                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

        # Column widths auto-adjustment (basic)
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            # Find max length in this column
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50) # Cap at 50

        # Freeze Panes (first row)
        ws.freeze_panes = "A2"

        # Prepare Response
        filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    finally:
        # Restore lxml
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']
