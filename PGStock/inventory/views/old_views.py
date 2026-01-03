import collections.abc
import sys

# Monkeypatch for Buffer in early Python 3.12 alpha versions
if not hasattr(collections.abc, 'Buffer'):
    class Buffer: pass
    collections.abc.Buffer = Buffer

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth import login, logout, authenticate

from django.contrib.auth.models import User, Group

from django.contrib.auth.decorators import login_required, user_passes_test

from django.contrib import messages

from django.core.paginator import Paginator

from django.core.exceptions import ValidationError

from django.db.models import F, Q, Sum, Count, Prefetch

from django.db.models.functions import Coalesce, TruncMonth

from django.db.models.deletion import ProtectedError
from django.db import transaction
from django.http import HttpResponse, JsonResponse

from django.utils import timezone

from django.utils import translation
from django.urls import reverse

from datetime import datetime, timedelta

from decimal import Decimal

from django_ratelimit.decorators import ratelimit



from ..models import (

    Category, Supplier, Client, Product, Inventory, PointOfSale,

    StockMovement, Invoice, InvoiceItem, Receipt, ReceiptItem, Payment, Settings,



    Quote, QuoteItem, PasswordResetCode, MonthlyProfitReport, Expense
)

from ..forms import (

    CustomUserCreationForm, CustomAuthenticationForm, 

    UserManageForm, UserProfileForm, SettingsForm,

    CategoryForm, SupplierForm, ClientForm, ProductForm,

    InventoryForm, StockMovementForm, InvoiceForm, InvoiceItemForm,

    ReceiptForm, ReceiptItemForm, QuoteForm, QuoteItemForm, PointOfSaleForm,
    ReplenishForm, PasswordResetRequestForm, PasswordResetVerifyForm, SetNewPasswordForm,
    ChangePasswordForm, QuoteItemFormSet, InvoiceItemFormSet
)

from django.core.mail import send_mail

from django.conf import settings

from django.utils.crypto import get_random_string

from ..models import UserProfile



# Import du système de permissions

from ..permissions import (

    admin_required, superuser_required, staff_required,

    is_admin, is_superuser_or_admin, is_staff_or_above,

    get_user_role, get_user_pos, filter_queryset_by_pos,
    can_modify_object, can_delete_object, validate_discount, can_view_finances
)



@ratelimit(key='ip', rate='10/m', block=True)
def user_login(request):

    """Vue de connexion"""

    if request.user.is_authenticated:

        # Vérifier si l'utilisateur a un rôle valide

        if is_staff_or_above(request.user):

            return redirect('inventory:dashboard')

        else:

            # Utilisateur authentifié mais sans rôle valide - le déconnecter

            logout(request)

            messages.error(request, "⛔ Votre compte n'a pas de rôle assigné. Veuillez contacter l'administrateur.")

    

    if request.method == 'POST':

        form = CustomAuthenticationForm(request, data=request.POST)

        if form.is_valid():

            username = form.cleaned_data.get('username')

            password = form.cleaned_data.get('password')

            user = authenticate(username=username, password=password)

            if user is not None:

                # Vérifier si l'utilisateur a un rôle valide avant de le connecter

                if is_staff_or_above(user):

                    login(request, user)

                    messages.success(request, f'Bienvenue {username}!')

                    return redirect('inventory:dashboard')

                else:

                    messages.error(request, "⛔ Votre compte n'a pas de rôle assigné. Veuillez contacter l'administrateur.")

    else:

        form = CustomAuthenticationForm()

    

    return render(request, 'inventory/auth/login.html', {'form': form})


@ratelimit(key='ip', rate='5/m', block=True)
def user_register(request):

    """Vue d'inscription"""

    if request.user.is_authenticated:

        return redirect('inventory:dashboard')

    

    if request.method == 'POST':

        form = CustomUserCreationForm(request.POST)

        if form.is_valid():

            user = form.save()

            login(request, user)

            messages.success(request, 'Compte créé avec succès!')

            return redirect('inventory:dashboard')

    else:

        form = CustomUserCreationForm()

    

    return render(request, 'inventory/auth/register.html', {'form': form})





@login_required

def user_logout(request):

    """Vue de déconnexion"""

    logout(request)

    # Render a dedicated logout page instead of redirecting immediately
    return render(request, 'inventory/auth/logout.html')





# ==================== PASSWORD RESET VIEWS ====================



def password_reset_request(request):

    if request.method == 'POST':

        form = PasswordResetRequestForm(request.POST)

        if form.is_valid():

            email = form.cleaned_data['email']

            # Use filter().first() to handle duplicate emails
            user = User.objects.filter(email=email).first()
            
            if not user:
                request.session['reset_email'] = email
                return redirect('inventory:password_reset_verify')

            

            # Generate code

            code = get_random_string(length=6, allowed_chars='0123456789')

            

            # Save code

            PasswordResetCode.objects.create(

                user=user,

                code=code,

                expires_at=timezone.now() + timedelta(minutes=15)

            )

            

            # Send email
            try:
                send_mail(
                    subject='Réinitialisation de votre mot de passe',
                    message=f'Votre code de vérification est : {code}. Ce code expire dans 15 minutes.',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
                messages.success(request, f'Un email contenant le code a été envoyé à {email}. Vérifiez votre dossier spam.')
            except Exception as e:
                # Log l'erreur pour le débogage mais ne pas bloquer l'utilisateur
                print(f"Erreur d'envoi d'email: {str(e)}")
                messages.error(request, 'Erreur lors de l\'envoi de l\'email. Contactez l\'administrateur si le problème persiste.')
            
            # Store email in session for next step
            request.session['reset_email'] = email
            return redirect('inventory:password_reset_verify')

    else:

        form = PasswordResetRequestForm()

    

    return render(request, 'inventory/password_reset_request.html', {'form': form})



def password_reset_verify(request):

    email = request.session.get('reset_email')

    if not email:

        return redirect('inventory:password_reset_request')

        

    if request.method == 'POST':

        form = PasswordResetVerifyForm(request.POST)

        if form.is_valid():

            code = form.cleaned_data['code']

            # Use filter().first() instead of get()
            user = User.objects.filter(email=email).first()
            
            if user:
                reset_code = PasswordResetCode.objects.filter(
                    user=user,
                    code=code,
                    used=False,
                    expires_at__gt=timezone.now()
                ).first()

                

                if reset_code:

                    reset_code.used = True

                    reset_code.save()

                    request.session['reset_verified'] = True

                    return redirect('inventory:password_reset_confirm')

                else:

                    form.add_error('code', 'Code invalide ou expiré.')

            else:
                form.add_error(None, 'Une erreur est survenue.')

    else:

        form = PasswordResetVerifyForm()

    

    return render(request, 'inventory/password_reset_verify.html', {'form': form, 'email': email})



def password_reset_confirm(request):

    email = request.session.get('reset_email')

    verified = request.session.get('reset_verified')

    

    if not email or not verified:

        return redirect('inventory:password_reset_request')

        

    if request.method == 'POST':

        form = SetNewPasswordForm(request.POST)

        if form.is_valid():

            new_password = form.cleaned_data['new_password']

            # Use filter().first() instead of get()
            user = User.objects.filter(email=email).first()

            if user:
                user.set_password(new_password)
                user.save()

            

            # Clean up session

            del request.session['reset_email']

            del request.session['reset_verified']

            

            messages.success(request, 'Votre mot de passe a été réinitialisé avec succès. Vous pouvez maintenant vous connecter.')

            return redirect('inventory:login')

    else:

        form = SetNewPasswordForm()

        

    return render(request, 'inventory/password_reset_confirm.html', {'form': form})







# ==================== USER MANAGEMENT VIEWS ====================



def is_admin(user):

    """Vérifie si l'utilisateur est un administrateur"""

    return user.is_superuser or user.groups.filter(name='Admin').exists()



@admin_required

def user_list(request):

    """Liste des utilisateurs (Admin seulement)"""

    all_users = User.objects.select_related('profile').prefetch_related('groups').all()

    

    # Calculate statistics for all users

    total_users = all_users.count()

    active_users_count = all_users.filter(is_active=True).count()

    inactive_users_count = all_users.filter(is_active=False).count()

    superuser_count = all_users.filter(is_superuser=True).count()

    

    users = all_users

    

    # Search functionality

    query = request.GET.get('q', '')

    if query:

        users = users.filter(

            Q(username__icontains=query) |

            Q(email__icontains=query) |

            Q(first_name__icontains=query) |

            Q(last_name__icontains=query)

        )

    

    # Status filter

    status_filter = request.GET.get('status', '')

    if status_filter == 'active':

        users = users.filter(is_active=True)

    elif status_filter == 'inactive':

        users = users.filter(is_active=False)

    elif status_filter == 'superuser':

        users = users.filter(is_superuser=True)

    

    users = users.order_by('username')

    # Pagination
    paginator = Paginator(users, 10)
    page = request.GET.get('page')
    users_paginated = paginator.get_page(page)
    
    return render(request, 'inventory/user/user_list.html', {
        'users': users_paginated,
        'page_obj': users_paginated,

        'query': query,

        'status_filter': status_filter,

        'total_users': total_users,

        'active_users_count': active_users_count,

        'inactive_users_count': inactive_users_count,

        'superuser_count': superuser_count,

    })





@admin_required

def user_create(request):

    """Créer un utilisateur (Admin seulement)"""

    if request.method == 'POST':

        user_form = UserManageForm(request.POST)

        profile_form = UserProfileForm(request.POST, request.FILES)

        

        if user_form.is_valid() and profile_form.is_valid():

            user = user_form.save()

            

            # Gérer le profil

            if not hasattr(user, 'profile'):

                UserProfile.objects.create(user=user)

            

            profile = user.profile

            if profile_form.cleaned_data.get('avatar'):

                profile.avatar = profile_form.cleaned_data['avatar']

                profile.save()

                

            messages.success(request, f"Utilisateur {user.username} créé avec succès!")

            return redirect('inventory:user_list')

    else:

        user_form = UserManageForm()

        profile_form = UserProfileForm()

    

    return render(request, 'inventory/user/user_form.html', {

        'user_form': user_form,

        'profile_form': profile_form,

        'title': 'Créer un utilisateur'

    })





@admin_required

def user_update(request, pk):

    """Modifier un utilisateur (Admin seulement)"""

    user = get_object_or_404(User, pk=pk)

    

    # Assurer que le profil existe

    if not hasattr(user, 'profile'):

        UserProfile.objects.create(user=user)

        

    if request.method == 'POST':

        user_form = UserManageForm(request.POST, instance=user)

        profile_form = UserProfileForm(request.POST, request.FILES, instance=user.profile)

        

        if user_form.is_valid() and profile_form.is_valid():

            user_form.save()

            profile_form.save()

            messages.success(request, f"Utilisateur {user.username} modifié avec succès!")

            return redirect('inventory:user_list')

    else:

        # Initialiser le champ group

        initial_group = user.groups.first()

        user_form = UserManageForm(instance=user, initial={'group': initial_group})

        profile_form = UserProfileForm(instance=user.profile)

    

    return render(request, 'inventory/user/user_form.html', {

        'user_form': user_form,

        'profile_form': profile_form,

        'title': f'Modifier {user.username}',

        'target_user': user

    })





@admin_required

def user_delete(request, pk):

    """Supprimer un utilisateur (Admin seulement)"""

    user = get_object_or_404(User, pk=pk)

    

    if user == request.user:

        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte!")

        return redirect('inventory:user_list')

        

    if request.method == 'POST':

        user.delete()

        messages.success(request, "Utilisateur supprimé avec succès!")

        return redirect('inventory:user_list')

    

    return render(request, 'inventory/user/user_confirm_delete.html', {'target_user': user})





# ==================== DASHBOARD ====================



@staff_required

def dashboard(request):

    """Tableau de bord principal"""

    # Statistiques générales

    total_products = Product.objects.count()

    total_categories = Category.objects.count()

    total_suppliers = Supplier.objects.count()

    total_clients = Client.objects.count()

    

    # Stock

    low_stock_count = Inventory.objects.filter(quantity__lte=F('reorder_level'), quantity__gt=0).count()

    out_of_stock_count = Inventory.objects.filter(quantity=0).count()

    total_stock_value = Product.objects.aggregate(

        total=Sum(F('selling_price') * F('inventory__quantity'))

    )['total'] or Decimal('0.00')

    

    # Bénéfice estimé
    total_estimated_profit = Decimal('0.00')
    if can_view_finances(request.user):
        total_estimated_profit = Product.objects.aggregate(
            total=Sum((F('selling_price') - F('purchase_price')) * F('inventory__quantity'))
        )['total'] or Decimal('0.00')
    
    # Ventes
    invoices_qs = Invoice.objects.all()
    invoices_qs = filter_queryset_by_pos(invoices_qs, request.user, 'point_of_sale')
    
    total_sales = invoices_qs.filter(status='paid').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')
    
    pending_orders = invoices_qs.filter(status='sent').count()
    
    # Mouvements récents
    movements_qs = StockMovement.objects.all()
    movements_qs = filter_queryset_by_pos(movements_qs, request.user, 'from_point_of_sale')
    recent_movements = movements_qs.select_related('product', 'user').order_by('-created_at')[:10]
    
    # Produits en stock faible
    inventory_qs = Inventory.objects.all()
    inventory_qs = filter_queryset_by_pos(inventory_qs, request.user, 'point_of_sale')
    
    low_stock_products = inventory_qs.filter(
        quantity__lte=F('reorder_level'),
        quantity__gt=0
    ).select_related('product')[:10]

    # Produits retournés (derniers 10)

    returned_products = StockMovement.objects.filter(

        movement_type='return'

    ).select_related('product').order_by('-created_at')[:10]

    

    invoice_stats = []

    statuses = [

        ('paid', 'Payées', 'bg-emerald-100 text-emerald-800 dark:bg-emerald-900 dark:text-emerald-200'),

        ('sent', 'Envoyées', 'bg-blue-100 text-blue-800 dark:bg-blue-900 dark:text-blue-200'),

        ('cancelled', 'Annulées', 'bg-red-100 text-red-800 dark:bg-red-900 dark:text-red-200'),

        ('draft', 'Brouillon', 'bg-gray-100 text-gray-800 dark:bg-gray-700 dark:text-gray-300')

    ]

    

    for status, label, css_class in statuses:
        stats = invoices_qs.filter(status=status).aggregate(
            count=Count('id'),
            total=Coalesce(Sum('total_amount'), Decimal('0.00'))
        )
        invoice_stats.append({
            'label': label,
            'status': status,
            'css_class': css_class,
            'count': stats['count'],
            'total': stats['total'] if can_view_finances(request.user) else Decimal('0.00')
        })

    

    # Produits défectueux (derniers 10)
    defective_products = movements_qs.filter(
        movement_type='defective'
    ).select_related('product').order_by('-created_at')[:10]
    
    # Activités récentes des ventes
    recent_sales_list = invoices_qs.filter(
        status__in=['sent', 'paid']
    ).select_related('client').order_by('-date_issued', '-created_at')
    
    # Mouvements récents
    recent_movements_list = movements_qs.select_related('product', 'user', 'from_point_of_sale').order_by('-created_at')

    # Produits en stock faible
    low_stock_products_list = inventory_qs.filter(
        quantity__lte=F('reorder_level'),
        quantity__gt=0
    ).select_related('product', 'point_of_sale')

    # Produits retournés
    returned_products_list = movements_qs.filter(
        movement_type='return'
    ).select_related('product').order_by('-created_at')

    # Pagination for all lists
    p_sales = Paginator(recent_sales_list, 5)

    p_movements = Paginator(recent_movements_list, 5)
    p_low_stock = Paginator(low_stock_products_list, 5)
    p_returns = Paginator(returned_products_list, 5)

    page_sales = request.GET.get('page_sales', 1)
    page_movements = request.GET.get('page_movements', 1)
    page_low_stock = request.GET.get('page_low_stock', 1)
    page_returns = request.GET.get('page_returns', 1)

    recent_sales = p_sales.get_page(page_sales)
    recent_movements = p_movements.get_page(page_movements)
    low_stock_products = p_low_stock.get_page(page_low_stock)
    returned_products = p_returns.get_page(page_returns)

    # Répartition du stock par point de vente
    stock_by_pos = PointOfSale.objects.annotate(
        total_items=Coalesce(Sum('inventory__quantity'), 0),
        total_value=Coalesce(Sum(F('inventory__quantity') * F('inventory__product__selling_price')), Decimal('0.00')),
        total_profit=Coalesce(Sum(F('inventory__quantity') * (F('inventory__product__selling_price') - F('inventory__product__purchase_price'))), Decimal('0.00'))
    ).filter(is_active=True)
    
    if not is_admin(request.user) and not request.user.is_superuser:
        user_pos = get_user_pos(request.user)
        if user_pos:
            stock_by_pos = stock_by_pos.filter(id=user_pos.id)
        else:
            stock_by_pos = PointOfSale.objects.none()

    # Répartition par catégorie
    category_data = Category.objects.annotate(
        count=Count('product')
    ).values('name', 'count').order_by('-count')

    top_selling_items = InvoiceItem.objects.values(
        'product__id', 'product__name', 'product__sku', 'product__image'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('unit_price'))
    ).order_by('-total_sold')[:5]

    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_suppliers': total_suppliers,
        'total_clients': total_clients,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'total_stock_value': total_stock_value if can_view_finances(request.user) else Decimal('0.00'),
        'total_sales': total_sales if can_view_finances(request.user) else Decimal('0.00'),
        'pending_orders': pending_orders,
        'recent_movements': recent_movements,
        'low_stock_products': low_stock_products,
        'returned_products': returned_products,
        'invoice_stats': invoice_stats,
        'defective_products': defective_products,
        'recent_sales': recent_sales,
        'stock_by_pos': stock_by_pos,
        'total_estimated_profit': total_estimated_profit if can_view_finances(request.user) else Decimal('0.00'),
        'category_data': list(category_data),
        'top_selling_items': top_selling_items,
    }
    
    return render(request, 'inventory/dashboard.html', context)



# ==================== CATEGORY VIEWS ====================



@staff_required

def category_list(request):

    """Liste des catégories"""

    query = request.GET.get('q', '')

    categories = Category.objects.annotate(product_count=Count('product'))

    

    if query:

        categories = categories.filter(

            Q(name__icontains=query) | Q(description__icontains=query)

        )

    

    return render(request, 'inventory/category/category_list.html', {

        'categories': categories,

        'query': query

    })





@superuser_required

def category_create(request):

    """Créer une catégorie"""

    if request.method == 'POST':

        form = CategoryForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(request, 'Catégorie créée avec succès!')

            return redirect('inventory:category_list')

    else:

        form = CategoryForm()

    

    return render(request, 'inventory/category/category_form.html', {'form': form})





@superuser_required

def category_update(request, pk):

    """Modifier une catégorie"""

    category = get_object_or_404(Category, pk=pk)

    

    if request.method == 'POST':

        form = CategoryForm(request.POST, instance=category)

        if form.is_valid():

            form.save()

            messages.success(request, 'Catégorie modifiée avec succès!')

            return redirect('inventory:category_list')

    else:

        form = CategoryForm(instance=category)

    

    return render(request, 'inventory/category/category_form.html', {

        'form': form,

        'category': category

    })





@admin_required

def category_delete(request, pk):

    """Supprimer une catégorie"""

    category = get_object_or_404(Category, pk=pk)

    

    if request.method == 'POST':

        try:

            category.delete()

            messages.success(request, 'Catégorie supprimée avec succès!')

        except ProtectedError:

            messages.error(request, "Impossible de supprimer cette catégorie car elle contient des produits.")

        return redirect('inventory:category_list')

    

    return render(request, 'inventory/category/category_confirm_delete.html', {

        'category': category

    })





# ==================== SUPPLIER VIEWS ====================



@staff_required

def supplier_list(request):

    """Liste des fournisseurs"""

    query = request.GET.get('q', '')

    suppliers = Supplier.objects.all()

    

    if query:

        suppliers = suppliers.filter(

            Q(name__icontains=query) | 

            Q(contact_person__icontains=query) |

            Q(email__icontains=query) |

            Q(phone__icontains=query)

        )

    

    paginator = Paginator(suppliers, 10)

    page = request.GET.get('page')

    suppliers = paginator.get_page(page)

    

    return render(request, 'inventory/supplier/supplier_list.html', {

        'suppliers': suppliers,

        'query': query

    })





@superuser_required

def supplier_create(request):

    """Créer un fournisseur"""

    if request.method == 'POST':

        form = SupplierForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(request, 'Fournisseur créé avec succès!')

            return redirect('inventory:supplier_list')

    else:

        form = SupplierForm()

    

    return render(request, 'inventory/supplier/supplier_form.html', {'form': form})





@superuser_required

def supplier_update(request, pk):

    """Modifier un fournisseur"""

    supplier = get_object_or_404(Supplier, pk=pk)

    

    if request.method == 'POST':

        form = SupplierForm(request.POST, instance=supplier)

        if form.is_valid():

            form.save()

            messages.success(request, 'Fournisseur modifié avec succès!')

            return redirect('inventory:supplier_list')

    else:

        form = SupplierForm(instance=supplier)

    

    return render(request, 'inventory/supplier/supplier_form.html', {

        'form': form,

        'supplier': supplier

    })





@staff_required

def supplier_detail(request, pk):

    """Détails d'un fournisseur"""

    supplier = get_object_or_404(Supplier, pk=pk)

    products = Product.objects.filter(supplier=supplier)

    receipts = Receipt.objects.filter(supplier=supplier).order_by('-date_received')[:10]

    

    return render(request, 'inventory/supplier/supplier_detail.html', {

        'supplier': supplier,

        'products': products,

        'receipts': receipts

    })





@admin_required

def supplier_delete(request, pk):

    """Supprimer un fournisseur"""

    supplier = get_object_or_404(Supplier, pk=pk)

    

    if request.method == 'POST':

        try:

            supplier.delete()

            messages.success(request, 'Fournisseur supprimé avec succès!')

        except ProtectedError:

            messages.error(request, "Impossible de supprimer ce fournisseur car il est lié à des produits ou des réceptions.")

        return redirect('inventory:supplier_list')

    

    return render(request, 'inventory/supplier/supplier_confirm_delete.html', {

        'supplier': supplier

    })





# ==================== CLIENT VIEWS ====================



@staff_required

def client_list(request):

    """Liste des clients"""

    query = request.GET.get('q', '')

    client_type = request.GET.get('type', '')

    

    clients = Client.objects.all()

    

    if query:

        clients = clients.filter(

            Q(name__icontains=query) | 

            Q(contact_person__icontains=query) |

            Q(email__icontains=query) |

            Q(phone__icontains=query)

        )

    

    if client_type:

        clients = clients.filter(client_type=client_type)

    

    paginator = Paginator(clients, 10)

    page = request.GET.get('page')

    clients = paginator.get_page(page)

    

    return render(request, 'inventory/client/client_list.html', {

        'clients': clients,

        'query': query,

        'client_type': client_type

    })





@staff_required

def client_create(request):

    """Créer un client"""

    if request.method == 'POST':

        form = ClientForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(request, 'Client créé avec succès!')

            return redirect('inventory:client_list')

    else:

        form = ClientForm()

    

    return render(request, 'inventory/client/client_form.html', {'form': form})





@superuser_required

def client_update(request, pk):

    """Modifier un client"""

    client = get_object_or_404(Client, pk=pk)

    

    if request.method == 'POST':

        form = ClientForm(request.POST, instance=client)

        if form.is_valid():

            form.save()

            messages.success(request, 'Client modifié avec succès!')

            return redirect('inventory:client_list')

    else:

        form = ClientForm(instance=client)

    

    return render(request, 'inventory/client/client_form.html', {

        'form': form,

        'client': client

    })





@staff_required

def client_detail(request, pk):

    """Détails d'un client"""

    client = get_object_or_404(Client, pk=pk)

    invoices = Invoice.objects.filter(client=client).order_by('-date_issued')[:10]

    

    return render(request, 'inventory/client/client_detail.html', {

        'client': client,

        'invoices': invoices

    })





@admin_required

def client_delete(request, pk):

    """Supprimer un client"""

    client = get_object_or_404(Client, pk=pk)

    

    if request.method == 'POST':

        try:

            client.delete()

            messages.success(request, 'Client supprimé avec succès!')

        except ProtectedError:

            messages.error(request, "Impossible de supprimer ce client car il a des factures associées.")

        return redirect('inventory:client_list')

    

    return render(request, 'inventory/client/client_confirm_delete.html', {

        'client': client

    })





# ==================== PRODUCT VIEWS ====================



@staff_required

def product_list(request):

    """Liste des produits"""

    query = request.GET.get('q', '')

    category_id = request.GET.get('category', '')

    

    products = Product.objects.select_related('category', 'supplier').prefetch_related(

        'inventory_set', 

        'inventory_set__point_of_sale'

    ).annotate(
        total_stock_annotated=Coalesce(Sum('inventory__quantity'), 0)
    ).all()
    
    # Filtrage par point de vente pour STAFF (Visibilité: Point de Vente uniquement)
    if not is_admin(request.user) and not request.user.is_superuser:
        user_pos = get_user_pos(request.user)
        if user_pos:
            products = products.filter(inventory__point_of_sale=user_pos).distinct()
        else:
            products = Product.objects.none()

    

    if query:

        products = products.filter(

            Q(name__icontains=query) | 

            Q(sku__icontains=query) | 

            Q(description__icontains=query)

        )

    

    if category_id:

        products = products.filter(category_id=category_id)

    

    # Order products for consistent pagination

    products = products.order_by('name')

    

    paginator = Paginator(products, 10)

    page = request.GET.get('page')

    products = paginator.get_page(page)

    

    categories = Category.objects.all()

    

    return render(request, 'inventory/product/product_list.html', {

        'products': products,

        'categories': categories,

        'query': query,

        'selected_category': category_id

    })





@superuser_required

def product_create(request):

    """Créer un produit"""

    from ..forms import ProductInventoryFormSet

    

    if request.method == 'POST':

        form = ProductForm(request.POST, request.FILES)

        if form.is_valid():

            product = form.save()

            

            # Gérer le formset d'inventaire

            formset = ProductInventoryFormSet(request.POST, instance=product)

            if formset.is_valid():

                formset.save()

                messages.success(request, 'Produit et inventaires créés avec succès!')

                return redirect('inventory:product_list')

            else:

                # Si le formset n'est pas valide, supprimer le produit et afficher les erreurs

                product.delete()

                messages.error(request, 'Erreur dans les données d\'inventaire.')

        else:

            # Si le formulaire produit n'est pas valide, initialiser quand même le formset pour l'affichage

            formset = ProductInventoryFormSet(request.POST)

    else:

        form = ProductForm()

        # Créer un formset vide mais prérempli avec tous les points de vente actifs

        formset = ProductInventoryFormSet(queryset=Inventory.objects.none())

        

        # Préremplir le formset avec tous les points de vente actifs

        active_pos = PointOfSale.objects.filter(is_active=True).order_by('name')

        initial_data = [{'point_of_sale': pos} for pos in active_pos]

        formset = ProductInventoryFormSet(initial=initial_data, queryset=Inventory.objects.none())

    

    return render(request, 'inventory/product/product_form.html', {

        'form': form,

        'inventory_formset': formset,

        'active_points_of_sale': PointOfSale.objects.filter(is_active=True).order_by('name'),

        'action': 'Créer'

    })





@superuser_required

def product_update(request, pk):

    """Modifier un produit"""

    from ..forms import ProductInventoryFormSet

    

    product = get_object_or_404(Product, pk=pk)

    

    if request.method == 'POST':

        form = ProductForm(request.POST, request.FILES, instance=product)

        formset = ProductInventoryFormSet(request.POST, instance=product)

        

        if form.is_valid() and formset.is_valid():

            form.save()

            formset.save()

            

            # Créer des inventaires pour les nouveaux points de vente

            active_pos = PointOfSale.objects.filter(is_active=True)

            for pos in active_pos:

                Inventory.objects.get_or_create(

                    product=product,

                    point_of_sale=pos,

                    defaults={'quantity': 0, 'reorder_level': 10}

                )

            

            messages.success(request, 'Produit et inventaires modifiés avec succès!')

            return redirect('inventory:product_detail', pk=pk)

    else:

        form = ProductForm(instance=product)

        

        # Charger les inventaires existants

        existing_inventories = product.inventory_set.all()

        

        # Créer des entrées pour les nouveaux points de vente (si nécessaire)

        active_pos = PointOfSale.objects.filter(is_active=True)

        for pos in active_pos:

            Inventory.objects.get_or_create(

                product=product,

                point_of_sale=pos,

                defaults={'quantity': 0, 'reorder_level': 10}

            )

        

        # Recharger le formset avec tous les inventaires

        formset = ProductInventoryFormSet(instance=product)

    

    return render(request, 'inventory/product/product_form.html', {

        'form': form,

        'product': product,

        'inventory_formset': formset,

        'active_points_of_sale': PointOfSale.objects.filter(is_active=True).order_by('name'),

        'action': 'Modifier'

    })





@staff_required
def product_detail(request, pk):
    """Détails d'un produit"""
    product = get_object_or_404(Product, pk=pk)
    # inventory = product.get_inventory()  # Deprecated single inventory
    inventories = product.inventory_set.select_related('point_of_sale').all()
    
    # Filtrage par point de vente pour STAFF
    if not is_admin(request.user) and not request.user.is_superuser:
        user_pos = get_user_pos(request.user)
        if user_pos:
            inventories = inventories.filter(point_of_sale=user_pos)
            
    # Pagination des mouvements
    movements_list = StockMovement.objects.filter(product=product).order_by('-created_at')
    
    # Filtrage par point de vente pour STAFF
    if not is_admin(request.user) and not request.user.is_superuser:
        user_pos = get_user_pos(request.user)
        if user_pos:
            movements_list = movements_list.filter(Q(from_point_of_sale=user_pos) | Q(to_point_of_sale=user_pos))
            
    paginator = Paginator(movements_list, 10) # 10 par page
    page = request.GET.get('page')
    movements = paginator.get_page(page)
    
    # Récupérer les paramètres pour la devise
    company_settings = Settings.objects.first()
    
    return render(request, 'inventory/product/product_detail.html', {
        'product': product,
        'inventories': inventories, # Pass full list
        'movements': movements,
        'company_settings': company_settings
    })





@admin_required

def product_delete(request, pk):

    """Supprimer un produit"""

    product = get_object_or_404(Product, pk=pk)

    

    if request.method == 'POST':

        try:

            product.delete()

            messages.success(request, 'Produit supprimé avec succès!')

        except ProtectedError:

            messages.error(request, "Impossible de supprimer ce produit car il est utilisé dans des factures ou des réceptions.")

        return redirect('inventory:product_list')

    

    return render(request, 'inventory/product/product_confirm_delete.html', {

        'product': product

    })





# ==================== INVENTORY VIEWS ====================



@staff_required
def inventory_list(request):
    """Liste de l'inventaire groupée par produit avec filtrage avancé"""
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    category_id = request.GET.get('category', '')
    pos_id = request.GET.get('pos', '')
    
    # On commence par filtrer les objets Inventory
    inventories = Inventory.objects.select_related('product', 'product__category', 'point_of_sale').all()
    
    # Filtrage par point de vente pour STAFF
    inventories = filter_queryset_by_pos(inventories, request.user, 'point_of_sale')
    
    # Recherche textuelle
    if query:
        inventories = inventories.filter(
            Q(product__name__icontains=query) | 
            Q(product__sku__icontains=query) |
            Q(location__icontains=query) |
            Q(product__description__icontains=query) |
            Q(product__category__name__icontains=query)
        )
    
    # Filtre par catégorie
    if category_id:
        inventories = inventories.filter(product__category_id=category_id)

    # Filtre par Point de Vente
    if pos_id:
        inventories = inventories.filter(point_of_sale_id=pos_id)

    # Groupement par produit
    from django.db.models import Sum, Count
    
    # On récupère les IDs des produits qui correspondent aux filtres
    product_ids = inventories.values_list('product_id', flat=True).distinct()
    
    # On prépare le queryset de base des produits
    products_qs = Product.objects.filter(id__in=product_ids).select_related('category').prefetch_related(
        Prefetch('inventory_set', queryset=inventories, to_attr='filtered_inventories')
    ).annotate(
        total_quantity=Sum('inventory__quantity'),
        pos_count=Count('inventory__point_of_sale', distinct=True)
    )

    # Filtre par statut (appliqué sur la quantité totale ou le statut spécifique)
    # Note: On garde la logique originale mais adaptée au groupement
    if status_filter:
        # Pour les filtres de statut, on doit parfois regarder au niveau global du produit
        if status_filter == 'low_stock':
            # Un produit est en stock faible s'il a au moins un inventaire en stock faible
            products_qs = products_qs.filter(inventory__quantity__lte=F('inventory__reorder_level'), inventory__quantity__gt=0).distinct()
        elif status_filter == 'out_of_stock':
            # Un produit est en rupture s'il n'a pas de stock du tout (ou 0 sur les filtres actuels)
            products_qs = products_qs.filter(total_quantity=0)
        elif status_filter == 'in_stock':
            products_qs = products_qs.filter(inventory__quantity__gt=F('inventory__reorder_level')).distinct()

    # Récupération des options pour les filtres
    categories = Category.objects.all().order_by('name')
    
    # Pour les points de vente
    if request.user.is_superuser or is_admin(request.user):
        points_of_sale = PointOfSale.objects.filter(is_active=True)
    else:
        user_pos = get_user_pos(request.user)
        points_of_sale = PointOfSale.objects.filter(id=user_pos.id) if user_pos else PointOfSale.objects.none()

    # Pagination sur les produits groupés
    paginator = Paginator(products_qs, 10)  # On peut augmenter un peu car c'est plus condensé
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/inventory/inventory_list.html', {
        'page_obj': page_obj,
        'search_query': query,
        'status_filter': status_filter,
        'category_filter': int(category_id) if category_id and category_id.isdigit() else None,
        'pos_filter': int(pos_id) if pos_id and pos_id.isdigit() else None,
        'categories': categories,
        'points_of_sale': points_of_sale,
    })


@staff_required
def export_inventory_excel(request):
    """Exporter l'inventaire en Excel"""
    from ..excel_utils import export_to_excel
    
    # Récupérer les filtres de la requête
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    category_id = request.GET.get('category', '')
    pos_id = request.GET.get('pos', '')
    
    inventories = Inventory.objects.select_related('product', 'product__category', 'point_of_sale').all()
    inventories = filter_queryset_by_pos(inventories, request.user, 'point_of_sale')
    
    if query:
        inventories = inventories.filter(
            Q(product__name__icontains=query) | 
            Q(product__sku__icontains=query) |
            Q(location__icontains=query) |
            Q(product__description__icontains=query) |
            Q(product__category__name__icontains=query)
        )
    
    if status_filter:
        if status_filter == 'low_stock':
            inventories = inventories.filter(quantity__lte=F('reorder_level'), quantity__gt=0)
        elif status_filter == 'out_of_stock':
            inventories = inventories.filter(quantity=0)
        elif status_filter == 'in_stock':
            inventories = inventories.filter(quantity__gt=F('reorder_level'))

    if category_id:
        inventories = inventories.filter(product__category_id=category_id)

    if pos_id:
        inventories = inventories.filter(point_of_sale_id=pos_id)
        
    headers = ['Produit', 'SKU', 'Catégorie', 'Qté Totale', 'Colis', 'Unités', 'Analyse', 'Seuil', 'Emplacement', 'Point de Vente', 'Statut']
    data = []
    
    for item in inventories:
        status_label = 'En stock'
        if item.quantity == 0:
            status_label = 'Rupture'
        elif item.quantity <= item.reorder_level:
            status_label = 'Stock faible'
            
        analysis_data = item.get_analysis_data()
            
        data.append([
            item.product.name,
            item.product.sku or '-',
            item.product.category.name if item.product.category else '-',
            item.quantity,
            analysis_data['colis'],
            analysis_data['unites'],
            analysis_data['analysis'],
            item.reorder_level,
            item.location or '-',
            item.point_of_sale.name,
            status_label
        ])
    
    return export_to_excel(headers, data, "Inventaire", "Inventaire")


@staff_required
def export_inventory_pdf(request):
    """Exporter l'inventaire en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
            
        # Récupérer les filtres
        query = request.GET.get('q', '')
        status_filter = request.GET.get('status', '')
        category_id = request.GET.get('category', '')
        pos_id = request.GET.get('pos', '')
        
        inventories = Inventory.objects.select_related('product', 'product__category', 'point_of_sale').all()
        inventories = filter_queryset_by_pos(inventories, request.user, 'point_of_sale')
        
        if query:
            inventories = inventories.filter(
                Q(product__name__icontains=query) | 
                Q(product__sku__icontains=query) |
                Q(location__icontains=query) |
                Q(product__description__icontains=query) |
                Q(product__category__name__icontains=query)
            )
        if status_filter:
            if status_filter == 'low_stock':
                inventories = inventories.filter(quantity__lte=F('reorder_level'), quantity__gt=0)
            elif status_filter == 'out_of_stock':
                inventories = inventories.filter(quantity=0)
            elif status_filter == 'in_stock':
                inventories = inventories.filter(quantity__gt=F('reorder_level'))
        if category_id:
            inventories = inventories.filter(product__category_id=category_id)
        if pos_id:
            inventories = inventories.filter(point_of_sale_id=pos_id)
            
        context = {
            'report_title': "Rapport d'Inventaire",
            'inventories': inventories[:500],
            'company_settings': Settings.objects.first(),
            'generated_at': timezone.now(),
        }
        
        template = get_template('inventory/reports_pdf/inventory_list_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Inventaire_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup: sys.modules['lxml'] = lxml_backup
        elif 'lxml' in sys.modules: del sys.modules['lxml']



@superuser_required

def inventory_update(request, pk):

    """Mise à jour d'un élément d'inventaire (AJAX ou standard)"""

    inventory = get_object_or_404(Inventory, pk=pk)

    

    if request.method == 'POST':

        location = request.POST.get('location')

        quantity = request.POST.get('quantity')

        reorder_level = request.POST.get('reorder_level')

        

        if location is not None:

            inventory.location = location

        

        if quantity:

            try:

                qty = int(quantity)

                if qty >= 0:

                    inventory.quantity = qty

            except ValueError:

                pass

                

        if reorder_level:

            try:

                level = int(reorder_level)

                if level >= 0:

                    inventory.reorder_level = level

            except ValueError:

                pass

        

        inventory.save()

        messages.success(request, f"Inventaire mis à jour pour {inventory.product.name}")

        

        return redirect('inventory:inventory_list')

    

    return redirect('inventory:inventory_list')





@staff_required
def inventory_detail(request, pk):
    """Détails de l'inventaire"""
    inventory = get_object_or_404(Inventory, pk=pk)
    movements = StockMovement.objects.filter(product=inventory.product).order_by('-created_at')[:20]
    
    # Récupérer les paramètres pour la devise
    company_settings = Settings.objects.first()
    
    return render(request, 'inventory/inventory/inventory_detail.html', {
        'inventory': inventory,
        'movements': movements,
        'company_settings': company_settings
    })





# ==================== STOCK MOVEMENT VIEWS ====================



@staff_required

def movement_list(request):

    """Liste des mouvements de stock"""

    # Récupérer les paramètres de recherche et filtrage
    search_query = request.GET.get('search', '')
    movement_type = request.GET.get('type', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    specific_date = request.GET.get('date', '')

    # Si une date précise est fournie, elle surcharge l'intervalle
    if specific_date:
        start_date = specific_date
        end_date = specific_date

    
    movements = StockMovement.objects.select_related('product', 'user', 'from_point_of_sale', 'to_point_of_sale').all()

    
    # Filtrage par point de vente pour STAFF
    movements = filter_queryset_by_pos(movements, request.user, 'from_point_of_sale')

    
    # Recherche textuelle (produit, SKU, référence, notes)
    if search_query:
        movements = movements.filter(
            Q(product__name__icontains=search_query) | 
            Q(product__sku__icontains=search_query) |
            Q(reference__icontains=search_query) |
            Q(notes__icontains=search_query)
        ).distinct()

    
    # Filtre par type de mouvement
    if movement_type:
        movements = movements.filter(movement_type=movement_type)

    
    # Filtre par utilisateur
    user_id = request.GET.get('user')
    if user_id:
        movements = movements.filter(user_id=user_id)

    # Filtre par produit
    product_id = request.GET.get('product')
    if product_id:
        movements = movements.filter(product_id=product_id)
        
    # Filtre par date de début
    if start_date:
        try:
            from datetime import datetime
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            movements = movements.filter(created_at__gte=start_datetime)
        except ValueError:
            pass  # Ignorer si le format de date est invalide

    # Filtre par date de fin
    if end_date:
        try:
            from datetime import datetime, timedelta
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # Inclure toute la journée de fin
            end_datetime = end_datetime + timedelta(days=1)
            movements = movements.filter(created_at__lt=end_datetime)
        except ValueError:
            pass  # Ignorer si le format de date est invalide

    # Tri par date décroissante
    movements = movements.order_by('-created_at')

    # Listes pour les filtres
    users = User.objects.filter(is_active=True).order_by('username')
    products = Product.objects.all().order_by('name')

    # Pagination
    paginator = Paginator(movements, 10)  # 10 items par page pour pagination moderne
    page = request.GET.get('page')
    movements = paginator.get_page(page)

    # Paramètres d'URL pour les liens d'export et de pagination
    # On exclue 'page' pour ne pas le dupliquer dans les liens de pagination
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    url_params = query_params.urlencode()

    return render(request, 'inventory/movement/movement_list.html', {
        'page_obj': movements,
        'search_query': search_query,
        'movement_type': movement_type,
        'start_date': start_date,
        'end_date': end_date,
        'specific_date': request.GET.get('date', ''), # Garder la valeur brute pour le template
        'users': users,
        'products': products,
        'selected_user': int(user_id) if user_id else None,
        'selected_product': int(product_id) if product_id else None,
        'url_params': url_params,
    })






@staff_required

def movement_create(request):

    """Créer un mouvement de stock"""

    if request.method == 'POST':

        form = StockMovementForm(request.POST)

        

        # Filtrer le QuerySet des points de vente pour empêcher de sélectionner un POS interdit

        if not is_admin(request.user):

            user_pos = get_user_pos(request.user)

            if user_pos:

                form.fields['from_point_of_sale'].queryset = PointOfSale.objects.filter(pk=user_pos.pk)

                # Auto-select si un seul choix

                form.initial['from_point_of_sale'] = user_pos

            else:

                form.fields['from_point_of_sale'].queryset = PointOfSale.objects.none()

        

        if form.is_valid():

            movement = form.save(commit=False)

            movement.user = request.user

            try:

                movement.save()

                

                # Feedback explicite sur le nouveau stock

                inventory = Inventory.objects.filter(product=movement.product, point_of_sale=movement.from_point_of_sale).first()

                new_qty = inventory.quantity if inventory else 0

                

                messages.success(

                    request, 

                    f'✅ {movement.get_movement_type_display()} enregistré pour {movement.product.name}. '

                    f'Nouveau stock au {movement.from_point_of_sale.code} : {new_qty}'

                )

                return redirect('inventory:movement_list')

            except ValidationError as e:

                # Capturer les erreurs de validation du modèle et les afficher

                messages.error(request, str(e))

    else:

        form = StockMovementForm()

        

        # Filtrer le QuerySet des points de vente pour l'affichage initial

        if not is_admin(request.user):

            user_pos = get_user_pos(request.user)

            if user_pos:

                form.fields['from_point_of_sale'].queryset = PointOfSale.objects.filter(pk=user_pos.pk)

                form.initial['from_point_of_sale'] = user_pos

                # Pour les transferts, on peut vouloir limiter 'to_point_of_sale' aussi, mais laissons ouvert pour l'instant

            else:

                form.fields['from_point_of_sale'].queryset = PointOfSale.objects.none()

                messages.warning(request, "⚠️ Aucun point de vente assigné à votre compte.")

    

    return render(request, 'inventory/movement/movement_form.html', {'form': form})





@staff_required

def movement_detail(request, pk):

    """Détails d'un mouvement de stock"""

    movement = get_object_or_404(StockMovement, pk=pk)

    

    return render(request, 'inventory/movement/movement_detail.html', {

        'movement': movement

    })





# ==================== INVOICE VIEWS ====================



@staff_required

def invoice_list(request):

    """Liste des factures"""

    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    invoices = Invoice.objects.select_related('client', 'created_by', 'point_of_sale').all()

    # Filtrage par point de vente pour STAFF
    invoices = filter_queryset_by_pos(invoices, request.user, 'point_of_sale')

    if query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=query) | 
            Q(client__name__icontains=query)
        )

    if status_filter:
        invoices = invoices.filter(status=status_filter)

    # Filtre par date de début
    if start_date:
        try:
            from datetime import datetime
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            invoices = invoices.filter(date_issued__gte=start_datetime)
        except ValueError:
            pass

    # Filtre par date de fin
    if end_date:
        try:
            from datetime import datetime, timedelta
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # Inclure toute la journée de fin si date_issued est un datetime, sinon juste la date
            invoices = invoices.filter(date_issued__lte=end_datetime)
        except ValueError:
            pass

    # Tri par date décroissante
    invoices = invoices.order_by('-date_issued', '-created_at')

    paginator = Paginator(invoices, 5)
    page = request.GET.get('page')
    invoices = paginator.get_page(page)

    return render(request, 'inventory/invoice/invoice_list.html', {
        'invoices': invoices,
        'query': query,
        'status_filter': status_filter,
        'start_date': start_date,
        'end_date': end_date
    })





@staff_required
@transaction.atomic
def invoice_create(request):
    """Créer une facture"""
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        formset = InvoiceItemFormSet(request.POST, form_kwargs={'user': request.user})
        
        if form.is_valid() and formset.is_valid():
            invoice = form.save(commit=False)
            invoice.created_by = request.user
            
            # Auto-assigner le point de vente depuis le profil utilisateur si non défini
            if not invoice.point_of_sale and hasattr(request.user, 'profile') and request.user.profile.point_of_sale:
                invoice.point_of_sale = request.user.profile.point_of_sale
            
            if not invoice.invoice_number:
                invoice.invoice_number = invoice.generate_invoice_number()
            invoice.save()
            
            items = formset.save(commit=False)
            for item in items:
                item.invoice = invoice
                item.save()
            for obj in formset.deleted_objects:
                obj.delete()
            
            invoice.calculate_totals()
            
            # Si la facture est créée directement avec le statut payé ou envoyé, déduire le stock
            if invoice.status in ['paid', 'sent']:
                try:
                    invoice.deduct_stock()
                    messages.success(request, 'Facture créée et stock déduit avec succès!')
                except ValueError as e:
                    messages.warning(request, f'Facture créée mais échec de la déduction du stock: {str(e)}')
            else:
                messages.success(request, 'Facture créée avec succès!')
                
            return redirect('inventory:invoice_detail', pk=invoice.pk)
    else:
        # Préremplir le point de vente depuis le profil utilisateur
        initial_data = {}
        if hasattr(request.user, 'profile') and request.user.profile.point_of_sale:
            initial_data['point_of_sale'] = request.user.profile.point_of_sale
        form = InvoiceForm(initial=initial_data)
        formset = InvoiceItemFormSet(form_kwargs={'user': request.user})
    
    return render(request, 'inventory/invoice/invoice_form.html', {
        'form': form,
        'item_formset': formset
    })





@staff_required

def invoice_detail(request, pk):

    """Détails d'une facture"""

    invoice = get_object_or_404(Invoice, pk=pk)

    items = invoice.invoiceitem_set.select_related('product').all()
    
    company_settings = Settings.objects.first()

    

    return render(request, 'inventory/invoice/invoice_detail.html', {

        'invoice': invoice,

        'items': items,
        
        'company_settings': company_settings

    })





@staff_required

def invoice_receipt(request, pk):

    """Ticket de caisse (format thermique)"""

    invoice = get_object_or_404(Invoice, pk=pk)

    items = invoice.invoiceitem_set.select_related('product').all()

    settings = Settings.objects.first()

    

    return render(request, 'inventory/invoice/receipt_small.html', {

        'invoice': invoice,

        'items': items,

        'settings': settings

    })





@superuser_required
@transaction.atomic
def invoice_update(request, pk):
    """Modifier une facture"""
    invoice = get_object_or_404(Invoice, pk=pk)
    
    # Vérifier les permissions de modification
    user_role = get_user_role(request.user)
    
    # STAFF ne peut modifier que les brouillons
    if user_role == 'STAFF' and invoice.status != 'draft':
        messages.error(request, "⛔ Vous ne pouvez modifier que les factures en brouillon.")
        return redirect('inventory:invoice_detail', pk=pk)
    
    # Vérifier le POS pour STAFF
    if user_role == 'STAFF':
        user_pos = get_user_pos(request.user)
        if user_pos and invoice.point_of_sale != user_pos:
            messages.error(request, "⛔ Vous ne pouvez modifier que les factures de votre point de vente.")
    
    old_status = invoice.status
    old_stock_deducted = invoice.stock_deducted
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        formset = InvoiceItemFormSet(request.POST, instance=invoice, form_kwargs={'user': request.user})
        
        if form.is_valid() and formset.is_valid():
            invoice = form.save()
            
            items = formset.save(commit=False)
            for item in items:
                item.invoice = invoice
                item.save()
            for obj in formset.deleted_objects:
                obj.delete()
            
            invoice.calculate_totals()
            
            # Si la facture est payée/envoyée ET le stock était déjà déduit,
            # il faut d'abord restaurer puis déduire à nouveau (pour gérer les modifications d'items)
            if invoice.status in ['paid', 'sent'] and old_stock_deducted:
                try:
                    # Restaurer d'abord le stock de l'ancienne facture
                    invoice.restore_stock()
                    # Puis déduire à nouveau avec les nouvelles valeurs
                    invoice.deduct_stock()
                    messages.success(request, 'Facture modifiée et stock mis à jour automatiquement!')
                except (ValueError, ValidationError) as e:
                    messages.warning(request, f'Facture modifiée mais le stock n\'a pas pu être mis à jour: {str(e)}')
            
            # Déduire le stock si le statut change à 'paid' ou 'sent' et que le stock n'a pas encore été déduit
            elif invoice.status in ['paid', 'sent'] and old_status not in ['paid', 'sent'] and not old_stock_deducted:
                try:
                    invoice.deduct_stock()
                    messages.success(request, 'Facture modifiée et stock déduit automatiquement!')
                except (ValueError, ValidationError) as e:
                    messages.warning(request, f'Facture modifiée mais le stock n\'a pas pu être déduit: {str(e)}')
            
            # Restaurer le stock si le statut change à 'cancelled' ou 'draft' et que le stock a été déduit
            elif invoice.status in ['cancelled', 'draft'] and old_stock_deducted:
                try:
                    invoice.restore_stock()
                    messages.success(request, 'Facture annulée et stock restauré automatiquement!')
                except (ValueError, ValidationError) as e:
                    messages.warning(request, f'Facture annulée mais le stock n\'a pas pu être restauré: {str(e)}')
            else:
                messages.success(request, 'Facture modifiée avec succès!')
            
            return redirect('inventory:invoice_detail', pk=pk)
    else:
        form = InvoiceForm(instance=invoice)
        formset = InvoiceItemFormSet(instance=invoice, form_kwargs={'user': request.user})
    
    return render(request, 'inventory/invoice/invoice_form.html', {
        'form': form,
        'invoice': invoice,
        'item_formset': formset
    })





@superuser_required

def invoice_delete(request, pk):

    """Supprimer une facture"""

    invoice = get_object_or_404(Invoice, pk=pk)

    

    if request.method == 'POST':

        invoice.delete()

        messages.success(request, 'Facture supprimée avec succès!')

        return redirect('inventory:invoice_list')

    

    return render(request, 'inventory/invoice/invoice_confirm_delete.html', {

        'invoice': invoice

    })





@staff_required

def invoice_add_item(request, pk):

    """Ajouter un article à une facture"""

    invoice = get_object_or_404(Invoice, pk=pk)

    

    if request.method == 'POST':

        form = InvoiceItemForm(request.POST, user=request.user)

        if form.is_valid():

            item = form.save(commit=False)

            item.invoice = invoice

            item.save()

            invoice.calculate_totals()

            

            # Si le stock a déjà été déduit pour cette facture, il faut déduire immédiatement pour ce nouvel article

            if invoice.stock_deducted:

                try:

                    StockMovement.objects.create(

                        product=item.product,

                        movement_type='exit',

                        quantity=item.quantity,

                        from_point_of_sale=invoice.point_of_sale,

                        reference=f"Facture {invoice.invoice_number} (Ajout)",

                        notes=f"Sortie automatique pour ajout article sur facture {invoice.invoice_number}",

                        user=request.user

                    )

                    messages.success(request, 'Article ajouté et stock déduit automatiquement!')

                except Exception as e:

                    messages.warning(request, f'Article ajouté mais erreur de stock: {str(e)}')

            else:

                messages.success(request, 'Article ajouté avec succès!')

                

            return redirect('inventory:invoice_detail', pk=pk)

    else:

        form = InvoiceItemForm(user=request.user)

    

    return render(request, 'inventory/invoice/invoice_add_item.html', {

        'form': form,

        'invoice': invoice

    })





@staff_required

def invoice_delete_item(request, pk, item_pk):

    """Supprimer un article d'une facture"""

    invoice = get_object_or_404(Invoice, pk=pk)

    item = get_object_or_404(InvoiceItem, pk=item_pk, invoice=invoice)

    

    if request.method == 'POST':

        # Si le stock a déjà été déduit, il faut le restaurer avant de supprimer l'article

        if invoice.stock_deducted:

            try:

                StockMovement.objects.create(

                    product=item.product,

                    movement_type='return',

                    quantity=item.quantity,

                    from_point_of_sale=invoice.point_of_sale,

                    reference=f"Facture {invoice.invoice_number} (Suppression)",

                    notes=f"Retour automatique suite à suppression article sur facture {invoice.invoice_number}",

                    user=request.user

                )

                stock_restored = True

            except Exception:

                stock_restored = False

        

        item.delete()

        invoice.calculate_totals()

        

        if invoice.stock_deducted:

            if stock_restored:

                messages.success(request, 'Article supprimé et stock restauré automatiquement!')

            else:

                messages.warning(request, 'Article supprimé mais le stock n\'a pas pu être restauré.')

        else:

            messages.success(request, 'Article supprimé avec succès!')

        

    return redirect('inventory:invoice_detail', pk=pk)






# ==================== QUOTE VIEWS ====================



@staff_required

def quote_list(request):

    """Liste des devis"""

    query = request.GET.get('q', '')

    status_filter = request.GET.get('status', '')

    

    quotes = Quote.objects.select_related('client', 'created_by').all()
    
    # Filtrage par point de vente pour STAFF (basé sur le créateur)
    if not is_admin(request.user) and not request.user.is_superuser:
        user_pos = get_user_pos(request.user)
        if user_pos:
            quotes = quotes.filter(created_by__profile__point_of_sale=user_pos)
        else:
            quotes = Quote.objects.none()

    

    if query:

        quotes = quotes.filter(

            Q(quote_number__icontains=query) | 

            Q(client__name__icontains=query)

        )

    

    if status_filter:

        quotes = quotes.filter(status=status_filter)

    

    paginator = Paginator(quotes, 10)

    page = request.GET.get('page')

    quotes = paginator.get_page(page)

    

    form = QuoteForm()
    return render(request, 'inventory/quote/quote_list.html', {
        'quotes': quotes,
        'query': query,
        'status_filter': status_filter,
        'form': form
    })





@staff_required
@transaction.atomic
def quote_create(request):
    """Créer un devis"""
    if request.method == 'POST':
        form = QuoteForm(request.POST)
        formset = QuoteItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            quote = form.save(commit=False)
            quote.created_by = request.user
            if not quote.quote_number:
                quote.quote_number = quote.generate_quote_number()
            quote.save()
            
            items = formset.save(commit=False)
            for item in items:
                item.quote = quote
                item.save()
            for obj in formset.deleted_objects:
                obj.delete()
                
            quote.calculate_totals()
            
            messages.success(request, 'Devis créé avec succès!')
            return redirect('inventory:quote_detail', pk=quote.pk)
    else:
        form = QuoteForm()
        formset = QuoteItemFormSet()
    
    return render(request, 'inventory/quote/quote_form.html', {
        'form': form,
        'item_formset': formset
    })


@staff_required
def quote_detail(request, pk):
    """Détails d'un devis"""
    quote = get_object_or_404(Quote, pk=pk)
    items = quote.quoteitem_set.select_related('product').all()
    
    return render(request, 'inventory/quote/quote_detail.html', {
        'quote': quote,
        'items': items
    })


@superuser_required
@transaction.atomic
def quote_update(request, pk):
    """Modifier un devis"""
    quote = get_object_or_404(Quote, pk=pk)
    
    if request.method == 'POST':
        form = QuoteForm(request.POST, instance=quote)
        formset = QuoteItemFormSet(request.POST, instance=quote)
        
        if form.is_valid() and formset.is_valid():
            quote = form.save()
            
            items = formset.save(commit=False)
            for item in items:
                item.quote = quote
                item.save()
            for obj in formset.deleted_objects:
                obj.delete()
                
            quote.calculate_totals()

            messages.success(request, 'Devis modifié avec succès!')
            return redirect('inventory:quote_detail', pk=pk)
    else:
        form = QuoteForm(instance=quote)
        formset = QuoteItemFormSet(instance=quote)
    
    return render(request, 'inventory/quote/quote_form.html', {
        'form': form,
        'quote': quote,
        'item_formset': formset
    })





@superuser_required

def quote_delete(request, pk):

    """Supprimer un devis"""

    quote = get_object_or_404(Quote, pk=pk)

    

    if request.method == 'POST':

        quote.delete()

        messages.success(request, 'Devis supprimé avec succès!')

        return redirect('inventory:quote_list')

    

    return render(request, 'inventory/quote/quote_confirm_delete.html', {

        'quote': quote

    })





@superuser_required
def quote_convert(request, pk):
    """Convertir un devis en facture"""
    quote = get_object_or_404(Quote, pk=pk)
    
    if quote.status == 'converted':
        messages.warning(request, 'Ce devis a déjà été converti en facture.')
        return redirect('inventory:quote_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            # Use the service layer for conversion
            from inventory.services import InvoiceService
            from inventory.services.base import ServiceException
            
            service = InvoiceService()
            
            # Get point of sale from user profile, or use first available
            point_of_sale = None
            if hasattr(request.user, 'profile') and request.user.profile.point_of_sale:
                point_of_sale = request.user.profile.point_of_sale
            else:
                # Fallback: use the first available point of sale
                point_of_sale = PointOfSale.objects.first()
            
            if not point_of_sale:
                messages.error(
                    request, 
                    'Aucun point de vente disponible dans le système. Veuillez en créer un d\'abord.'
                )
                return redirect('inventory:quote_detail', pk=pk)
            
            # Convert quote to invoice using the service
            invoice = service.convert_quote_to_invoice(
                quote=quote,
                user=request.user,
                point_of_sale=point_of_sale
            )
            
            messages.success(
                request,
                f'Devis converti en facture {invoice.invoice_number} avec succès!'
            )
            return redirect('inventory:invoice_detail', pk=invoice.pk)
            
        except ServiceException as e:
            messages.error(request, str(e))
            return redirect('inventory:quote_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Erreur lors de la conversion: {str(e)}')
            return redirect('inventory:quote_detail', pk=pk)
    
    return render(request, 'inventory/quote/quote_confirm_convert.html', {
        'quote': quote
    })






@login_required

def quote_add_item(request, pk):

    """Ajouter un article à un devis"""

    quote = get_object_or_404(Quote, pk=pk)

    

    if request.method == 'POST':

        form = QuoteItemForm(request.POST)

        if form.is_valid():

            item = form.save(commit=False)

            item.quote = quote

            item.save()

            quote.calculate_totals()

            

            messages.success(request, 'Article ajouté avec succès!')

            return redirect('inventory:quote_detail', pk=pk)

    else:

        form = QuoteItemForm()

    

    return render(request, 'inventory/quote/quote_add_item.html', {

        'form': form,

        'quote': quote

    })





@login_required

def quote_delete_item(request, pk, item_pk):

    """Supprimer un article d'un devis"""

    quote = get_object_or_404(Quote, pk=pk)

    item = get_object_or_404(QuoteItem, pk=item_pk, quote=quote)

    

    if request.method == 'POST':

        item.delete()

        quote.calculate_totals()

        messages.success(request, 'Article supprimé avec succès!')

        

    return redirect('inventory:quote_detail', pk=pk)





# ==================== API VIEWS FOR CHARTS ====================



@login_required

def api_stock_evolution(request):

    """API pour l'évolution des mouvements de stock"""

    days = int(request.GET.get('days', 30))

    end_date = timezone.now()

    start_date = end_date - timedelta(days=days)

    

    movements = StockMovement.objects.filter(

        created_at__gte=start_date

    ).values('created_at__date', 'movement_type').annotate(

        total=Sum('quantity')

    ).order_by('created_at__date')

    

    data = {

        'labels': [],

        'entries': [],

        'exits': []

    }

    

    # Process data

    date_dict = {}

    for movement in movements:

        date_str = movement['created_at__date'].strftime('%Y-%m-%d')

        if date_str not in date_dict:

            date_dict[date_str] = {'entry': 0, 'exit': 0}

        

        if movement['movement_type'] in ['entry', 'return']:

            date_dict[date_str]['entry'] += movement['total']

        elif movement['movement_type'] in ['exit', 'transfer', 'defective']:

            date_dict[date_str]['exit'] += movement['total']

    

    for date_str in sorted(date_dict.keys()):

        data['labels'].append(date_str)

        data['entries'].append(date_dict[date_str]['entry'])

        data['exits'].append(date_dict[date_str]['exit'])

        

    return JsonResponse(data)





@login_required

def api_category_distribution(request):

    """API pour la distribution des produits par catégorie"""

    categories = Category.objects.annotate(

        total_quantity=Coalesce(Sum('product__inventory__quantity'), 0)

    ).values('name', 'total_quantity')

    

    data = {

        'labels': [cat['name'] for cat in categories],

        'data': [cat['total_quantity'] for cat in categories]

    }

    

    return JsonResponse(data)





@login_required

def api_monthly_revenue(request):

    """API pour l'évolution mensuelle des revenus sur 12 mois"""

    twelve_months_ago = timezone.now() - timedelta(days=365)

    

    revenue_data = Invoice.objects.filter(

        status='paid',

        date_issued__gte=twelve_months_ago

    ).annotate(

        month=TruncMonth('date_issued')

    ).values('month').annotate(

        total=Sum('total_amount')

    ).order_by('month')

    

    data = {

        'labels': [item['month'].strftime('%b %Y') for item in revenue_data],

        'data': [float(item['total']) for item in revenue_data]

    }

    

    return JsonResponse(data)





@login_required

def api_product_sales_type(request):

    """API pour comparer les ventes en gros vs détail des top produits"""

    # Top 5 produits par quantité totale vendue

    top_products = InvoiceItem.objects.values('product__name').annotate(

        total_qty=Sum('quantity')

    ).order_by('-total_qty')[:5]

    

    product_names = [p['product__name'] for p in top_products]

    

    retail_data = []

    wholesale_data = []

    

    for name in product_names:

        retail_qty = InvoiceItem.objects.filter(

            product__name=name, 

            is_wholesale=False,

            invoice__status='paid'

        ).aggregate(total=Coalesce(Sum('quantity'), 0))['total']

        

        wholesale_qty = InvoiceItem.objects.filter(

            product__name=name, 

            is_wholesale=True,

            invoice__status='paid'

        ).aggregate(total=Coalesce(Sum('quantity'), 0))['total']

        

        retail_data.append(int(retail_qty))

        wholesale_data.append(int(wholesale_qty))

        

    data = {

        'labels': product_names,

        'retail': retail_data,

        'wholesale': wholesale_data

    }

    

    return JsonResponse(data)








@login_required

def api_product_info(request, pk):

    """API pour récupérer les informations d'un produit"""

    try:

        product = Product.objects.get(pk=pk)

        

        # Récupérer le point de vente depuis la requête (pour afficher le stock disponible)

        pos_id = request.GET.get('pos_id')

        stock_info = {}

        

        if pos_id:

            try:

                pos = PointOfSale.objects.get(pk=pos_id)

                inventory = Inventory.objects.filter(product=product, point_of_sale=pos).first()

                if inventory:

                    stock_info = {

                        'quantity': inventory.quantity,

                        'location': inventory.location,

                        'status': inventory.get_status(),

                        'status_display': inventory.get_status_display(),

                        'reorder_level': inventory.reorder_level

                    }

                else:

                    stock_info = {

                        'quantity': 0,

                        'location': '',

                        'status': 'out_of_stock',

                        'status_display': 'Rupture de stock',

                        'reorder_level': 0

                    }

            except PointOfSale.DoesNotExist:

                pass

        

        # Calculer le stock total sur tous les points de vente

        total_stock = product.get_total_stock_quantity()

        

        data = {

            'id': product.id,

            'name': product.name,

            'sku': product.sku,

            'purchase_price': str(product.purchase_price),

            'margin': str(product.margin),

            'selling_price': str(product.selling_price),
            'units_per_box': product.units_per_box,
            'wholesale_purchase_price': str(product.wholesale_purchase_price),
            'wholesale_margin': str(product.wholesale_margin),
            'wholesale_selling_price': str(product.wholesale_selling_price),

            'total_stock': total_stock,

            'stock_info': stock_info,

            'category': product.category.name if product.category else '',

            'supplier': product.supplier.name if product.supplier else ''

        }

        

        return JsonResponse(data)

    except Product.DoesNotExist:

        return JsonResponse({'error': 'Produit non trouvé'}, status=404)





@superuser_required

def reports_view(request):

    """Vue des rapports avec filtrage par date"""

    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    specific_date = request.GET.get('date')

    if specific_date:
        start_date = specific_date
        end_date = specific_date

    

    # Base querysets

    invoices = Invoice.objects.all()

    movements = StockMovement.objects.all()

    

    if start_date:

        invoices = invoices.filter(date_issued__gte=start_date)

        movements = movements.filter(created_at__date__gte=start_date)

    

    if end_date:

        invoices = invoices.filter(date_issued__lte=end_date)

        movements = movements.filter(created_at__date__lte=end_date)

    

    # Stock stats (Always current state, not filtered by date)

    total_stock_value = Product.objects.aggregate(

        total=Sum(F('selling_price') * F('inventory__quantity'))

    )['total'] or Decimal('0.00')

    

    total_products = Product.objects.count()

    

    # Get products where total stock across all POS is 0

    # This matches the logic in Product.get_stock_status()

    products_with_zero_stock = Product.objects.annotate(

        total_quantity=Coalesce(Sum('inventory__quantity'), 0)

    ).filter(total_quantity=0).select_related('category', 'supplier')

    

    out_of_stock_count = products_with_zero_stock.count()

    

    # For display, we'll show these products with their first POS (or create a list showing all POS)

    # We'll create a list of products with their inventory details

    out_of_stock_items = []

    for product in products_with_zero_stock[:10]:

        # Get all inventory records for this product (even if they're all 0)

        inventories = product.inventory_set.select_related('point_of_sale').all()

        if inventories:

            # Add the first inventory record for display

            out_of_stock_items.append({

                'product': product,

                'point_of_sale': inventories[0].point_of_sale,

                'all_pos': [inv.point_of_sale for inv in inventories]

            })

        else:

            # Product has no inventory records at all

            out_of_stock_items.append({

                'product': product,

                'point_of_sale': None,

                'all_pos': []

            })

    

    low_stock_count = Inventory.objects.filter(quantity__lte=F('reorder_level'), quantity__gt=0).count()

    

    # Sales stats (Filtered)

    total_sales = invoices.filter(status='paid').aggregate(

        total=Sum('total_amount')

    )['total'] or Decimal('0.00')

    

    pending_invoices = invoices.filter(status='sent')

    pending_invoices_count = pending_invoices.count()

    pending_sales_amount = pending_invoices.aggregate(

        total=Sum('total_amount')

    )['total'] or Decimal('0.00')

    

    # Recent sales (NOT filtered by date - always show most recent)

    recent_sales = Invoice.objects.filter(status='paid').select_related('client').order_by('-date_issued')[:10]

    

    # Top products by sales (Filtered)

    # Note: We need to filter InvoiceItems based on the filtered Invoices

    top_products = InvoiceItem.objects.filter(invoice__in=invoices.filter(status='paid')).values('product__name').annotate(

        total_quantity=Sum('quantity'),

        total_revenue=Sum('total')

    ).order_by('-total_revenue')[:5]

    

    # Recent stock movements (NOT filtered by date - always show most recent)

    recent_movements = StockMovement.objects.select_related('product', 'user').order_by('-created_at')[:10]

    

    # --- DAILY STATISTICS ---

    

    # 1. Global Daily Stats

    daily_sales_qs = invoices.filter(status='paid').values('date_issued').annotate(

        total=Coalesce(Sum('total_amount'), Decimal('0.00'))

    ).order_by('-date_issued')

    

    daily_purchases_qs = Receipt.objects.filter(status='validated', date_received__gte=start_date if start_date else '1900-01-01').values('date_received').annotate(

        total=Coalesce(Sum('total_amount'), Decimal('0.00'))

    ).order_by('-date_received')

    

    if end_date:

        daily_purchases_qs = daily_purchases_qs.filter(date_received__lte=end_date)

        

    # Merge into a single dictionary by date

    daily_stats = {}

    

    for sale in daily_sales_qs:

        date_str = sale['date_issued'].strftime('%Y-%m-%d')

        if date_str not in daily_stats:

            daily_stats[date_str] = {'date': sale['date_issued'], 'sales': Decimal('0.00'), 'purchases': Decimal('0.00'), 'profit': Decimal('0.00')}

        daily_stats[date_str]['sales'] = sale['total']

        

    for purchase in daily_purchases_qs:

        date_str = purchase['date_received'].strftime('%Y-%m-%d')

        if date_str not in daily_stats:

            daily_stats[date_str] = {'date': purchase['date_received'], 'sales': Decimal('0.00'), 'purchases': Decimal('0.00'), 'profit': Decimal('0.00')}

        daily_stats[date_str]['purchases'] = purchase['total']

    

    # Calculate profit for each day

    for date_str, stats in daily_stats.items():

        stats['profit'] = stats['sales'] - stats['purchases']

        

    # Convert to sorted list

    daily_stats_list = sorted(daily_stats.values(), key=lambda x: x['date'], reverse=True)

    

    # 2. Per POS Daily Stats

    pos_stats = []

    points_of_sale = PointOfSale.objects.filter(is_active=True)

    

    for pos in points_of_sale:

        # Sales for this POS

        pos_sales = invoices.filter(status='paid', point_of_sale=pos).values('date_issued').annotate(

            total=Coalesce(Sum('total_amount'), Decimal('0.00'))

        )

        

        # Purchases for this POS

        pos_receipts = Receipt.objects.filter(

            status='validated', 

            point_of_sale=pos,

            date_received__gte=start_date if start_date else '1900-01-01'

        )

        if end_date:

            pos_receipts = pos_receipts.filter(date_received__lte=end_date)

            

        pos_purchases = pos_receipts.values('date_received').annotate(

            total=Coalesce(Sum('total_amount'), Decimal('0.00'))

        )

        

        # Merge for this POS

        pos_daily = {}

        total_sales_pos = Decimal('0.00')

        total_purchases_pos = Decimal('0.00')

        

        for sale in pos_sales:

            date_str = sale['date_issued'].strftime('%Y-%m-%d')

            if date_str not in pos_daily:

                pos_daily[date_str] = {'date': sale['date_issued'], 'sales': Decimal('0.00'), 'purchases': Decimal('0.00')}

            pos_daily[date_str]['sales'] = sale['total']

            total_sales_pos += sale['total']

            

        for purchase in pos_purchases:

            date_str = purchase['date_received'].strftime('%Y-%m-%d')

            if date_str not in pos_daily:

                pos_daily[date_str] = {'date': purchase['date_received'], 'sales': Decimal('0.00'), 'purchases': Decimal('0.00')}

            pos_daily[date_str]['purchases'] = purchase['total']

            total_purchases_pos += purchase['total']

            

        if pos_daily: # Only add if there is data

            pos_stats.append({

                'pos': pos,

                'daily_data': sorted(pos_daily.values(), key=lambda x: x['date'], reverse=True),

                'total_sales': total_sales_pos,

                'total_purchases': total_purchases_pos

            })

            

    # Calculate Grand Totals (Purchases) - Sales is already 'total_sales' variable

    total_purchases = Receipt.objects.filter(

        status='validated',

        date_received__gte=start_date if start_date else '1900-01-01'

    )

    if end_date:

        total_purchases = total_purchases.filter(date_received__lte=end_date)

        

    total_purchases_amount = total_purchases.aggregate(

        total=Coalesce(Sum('total_amount'), Decimal('0.00'))

    )['total']

    

    total_gross_profit = total_sales - total_purchases_amount

    

    context = {

        'total_stock_value': total_stock_value,

        'total_products': total_products,

        'out_of_stock_count': out_of_stock_count,

        'low_stock_count': low_stock_count,

        'total_sales': total_sales,

        'total_purchases': total_purchases_amount,

        'total_gross_profit': total_gross_profit,

        'pending_invoices_count': pending_invoices_count,

        'pending_sales_amount': pending_sales_amount,

        'recent_sales': recent_sales,

        'top_products': top_products,

        'recent_movements': recent_movements,

        'start_date': start_date,

        'end_date': end_date,

        'daily_stats': daily_stats_list,

        'pos_stats': pos_stats,

        'out_of_stock_items': out_of_stock_items,

    }

    

    return render(request, 'inventory/reports.html', context)





# ==================== PAYMENT VIEWS ====================



@staff_required
def payment_list(request):
    """Liste des paiements et suivi des factures"""
    payments = Payment.objects.select_related('invoice', 'invoice__client').all()
    
    # Filtrage par point de vente pour STAFF
    payments = filter_queryset_by_pos(payments, request.user, 'invoice__point_of_sale')

    # Factures avec solde à payer
    invoices_qs = Invoice.objects.filter(status__in=['sent', 'partial']).select_related('client')
    invoices_qs = filter_queryset_by_pos(invoices_qs, request.user, 'point_of_sale')
    
    unpaid_invoices = [
        inv for inv in invoices_qs 
        if inv.get_remaining_amount() > 0
    ]

    

    # Stats

    total_collected = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_outstanding = sum(inv.get_remaining_amount() for inv in unpaid_invoices)

    

    # Pagination pour l'historique des transactions
    paginator = Paginator(payments, 10)
    page = request.GET.get('page')
    payments_paginated = paginator.get_page(page)
    
    context = {
        'payments': payments_paginated,
        'page_obj': payments_paginated, # Pour le composant de pagination
        'unpaid_invoices': unpaid_invoices,
        'total_collected': total_collected,
        'total_outstanding': total_outstanding,
    }



    return render(request, 'inventory/payment/payment_list.html', context)



@login_required
def payment_create(request, pk):
    """Enregistrer un paiement pour une facture (Supporte AJAX/Modal)"""
    invoice = get_object_or_404(Invoice, pk=pk)
    
    # Gestion de la requête AJAX pour afficher le modal (GET)
    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'inventory/payment/payment_modal.html', {'invoice': invoice})

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Récupération et nettoyage des données
                amount_str = request.POST.get('amount', '').strip().replace(',', '.')
                payment_date_str = request.POST.get('payment_date')
                payment_method = request.POST.get('payment_method')
                reference = request.POST.get('reference', '').strip()
                notes = request.POST.get('notes', '').strip()
                
                # 2. Validations
                if not amount_str:
                    raise ValueError("Le montant est requis.")
                
                try:
                    amount = Decimal(amount_str)
                except Exception:
                    raise ValueError("Le format du montant est invalide.")
                
                if amount <= 0:
                    raise ValueError("Le montant doit être strictement positif.")
                
                remaining = invoice.get_remaining_amount()
                if amount > remaining:
                    # Tolérance minime pour les erreurs d'arrondi ou UX
                    if amount - remaining > Decimal('0.01'):
                        raise ValueError(f"Le montant ({amount}) ne peut pas dépasser le reste à payer ({remaining}).")
                
                valid_methods = [m[0] for m in Payment.PAYMENT_METHODS]
                if payment_method not in valid_methods:
                    raise ValueError("Mode de paiement invalide.")
                
                if not payment_date_str:
                     payment_date = timezone.now().date()
                else:
                    try:
                         # Tente de parser, sinon laisse Django/DB gérer ou default
                         payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                    except ValueError:
                         raise ValueError("Date de paiement invalide (format attendu: YYYY-MM-DD).")
                
                # 3. Création du paiement
                payment = Payment.objects.create(
                    invoice=invoice,
                    amount=amount,
                    payment_date=payment_date,
                    payment_method=payment_method,
                    reference=reference,
                    notes=notes,
                    created_by=request.user
                )
                
                # 4. Feedback
                success_msg = f"Paiement de {amount} enregistré avec succès."
                
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({
                        'status': 'success', 
                        'message': success_msg,
                        'redirect_url': request.build_absolute_uri(reverse('inventory:invoice_detail', args=[pk]))
                    })
                
                messages.success(request, success_msg)
                return redirect('inventory:invoice_detail', pk=pk)
            
        except ValueError as e:
            error_msg = str(e)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                 return JsonResponse({'status': 'error', 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('inventory:invoice_detail', pk=pk)
            
        except Exception as e:
            # Log l'erreur réelle côté serveur pour débogage si besoin
            # logger.error(f"Erreur paiement: {e}") 
            error_msg = f"Une erreur technique est survenue lors de l'enregistrement."
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                 return JsonResponse({'status': 'error', 'message': f"{error_msg} ({str(e)})"}) # Dev mode: show error
            messages.error(request, error_msg)
            return redirect('inventory:invoice_detail', pk=pk)

    # Fallback pour GET non-AJAX
    return redirect('inventory:invoice_detail', pk=pk)







# ==================== SETTINGS VIEWS ====================



@login_required

def settings_view(request):

    """Vue des paramètres"""

    # Get or create the singleton settings object

    settings_obj = Settings.objects.first()

    if not settings_obj:

        settings_obj = Settings.objects.create()

    

    if request.method == 'POST':

        settings_form = SettingsForm(request.POST, request.FILES, instance=settings_obj)

        if settings_form.is_valid():

            saved_settings = settings_form.save()

            # Activer immédiatement la nouvelle langue

            if saved_settings.language:

                translation.activate(saved_settings.language)

                request.LANGUAGE_CODE = saved_settings.language

            messages.success(request, 'Paramètres mis à jour avec succès!')

            return redirect('inventory:settings')

    else:

        settings_form = SettingsForm(instance=settings_obj)



    context = {

        'total_products': Product.objects.count(),

        'total_categories': Category.objects.count(),

        'total_movements': StockMovement.objects.count(),

        'settings_form': settings_form,

        'password_form': ChangePasswordForm(user=request.user),

    }

    return render(request, 'inventory/settings.html', context)





@login_required

def update_profile(request):

    """Mise à jour du profil utilisateur"""

    if request.method == 'POST':

        user = request.user

        user.first_name = request.POST.get('first_name', '')

        user.last_name = request.POST.get('last_name', '')

        user.email = request.POST.get('email', '')

        user.save()

        messages.success(request, 'Profil mis à jour avec succès!')

        return redirect('inventory:settings')

    

    return redirect('inventory:settings')





@login_required

def change_password(request):

    """Changement de mot de passe"""

    if request.method == 'POST':

        form = ChangePasswordForm(user=request.user, data=request.POST)

        if form.is_valid():

            from django.contrib.auth import update_session_auth_hash

            

            new_password = form.cleaned_data['new_password']

            request.user.set_password(new_password)

            request.user.save()

            

            # Garder l'utilisateur connecté après le changement

            update_session_auth_hash(request, request.user)

            

            messages.success(request, '✅ Mot de passe changé avec succès!')

            return redirect('inventory:settings')

        else:

            # Si le formulaire n'est pas valide, afficher les erreurs

            for field, errors in form.errors.items():

                for error in errors:

                    messages.error(request, error)

            return redirect('inventory:settings')

    

    return redirect('inventory:settings')





@login_required
def export_reports_excel(request):
    """Exporter les rapports en Excel"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Design tokens
        PRIMARY_BLUE = '4472C4'
        WHITE = 'FFFFFF'
        LIGHT_GRAY = 'F4F6FA'
        BORDER_COLOR = 'D1D5DB'

        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        specific_date = request.GET.get('date')

        if specific_date:
            start_date = specific_date
            end_date = specific_date
        
        if start_date in [None, '', 'None']: start_date = None
        if end_date in [None, '', 'None']: end_date = None
        
        invoices = Invoice.objects.select_related('client').all()
        movements = StockMovement.objects.select_related('product', 'user').all()
        inventories = Inventory.objects.select_related('product', 'product__category').all()
        
        if start_date:
            invoices = invoices.filter(date_issued__gte=start_date)
            movements = movements.filter(created_at__date__gte=start_date)
        if end_date:
            invoices = invoices.filter(date_issued__lte=end_date)
            movements = movements.filter(created_at__date__lte=end_date)
            
        wb = openpyxl.Workbook()
        
        def apply_modern_style(ws, monetary_cols=None):
            # Header style
            header_fill = PatternFill(start_color=PRIMARY_BLUE, end_color=PRIMARY_BLUE, fill_type="solid")
            header_font = Font(bold=True, color=WHITE)
            border = Border(left=Side(style='thin', color=BORDER_COLOR), 
                            right=Side(style='thin', color=BORDER_COLOR), 
                            top=Side(style='thin', color=BORDER_COLOR), 
                            bottom=Side(style='thin', color=BORDER_COLOR))
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Alternating row colors and column adjustment
            row_count = 0
            for row in ws.iter_rows(min_row=2):
                row_count += 1
                fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if row_count % 2 == 0 else None
                for cell in row:
                    if fill: cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Apply currency format to specified columns
                    if monetary_cols and cell.column in monetary_cols:
                        cell.number_format = '#,##0.00'

            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 4, 50)

        # Ventes
        ws_sales = wb.active
        ws_sales.title = "Ventes"
        ws_sales.append(['Date', 'Facture', 'Client', 'Statut', 'Montant Total', 'Payé', 'Reste'])
        for inv in invoices:
            ws_sales.append([inv.date_issued.strftime('%d/%m/%Y'), inv.invoice_number, inv.client.name, 
                             inv.get_status_display(), float(inv.total_amount), 
                             float(inv.get_amount_paid()), float(inv.get_remaining_amount())])
        apply_modern_style(ws_sales, monetary_cols=[5, 6, 7])

        # Etat du Stock
        ws_stock = wb.create_sheet("Etat du Stock")
        ws_stock.append(['Produit', 'SKU', 'Catégorie', 'Quantité', 'Prix Achat', 'Prix Vente', 'Valeur Totale'])
        for inv in inventories:
            ws_stock.append([inv.product.name, inv.product.sku or '-', inv.product.category.name if inv.product.category else '-',
                             inv.quantity, float(inv.product.purchase_price or 0), float(inv.product.selling_price or 0),
                             float(inv.quantity * (inv.product.selling_price or 0))])
        apply_modern_style(ws_stock, monetary_cols=[5, 6, 7])

        # Mouvements
        ws_movements = wb.create_sheet("Mouvements de Stock")
        ws_movements.append(['Date', 'Produit', 'Type', 'Quantité', 'Référence', 'Utilisateur', 'Notes'])
        for mov in movements:
            ws_movements.append([mov.created_at.strftime('%d/%m/%Y %H:%M'), mov.product.name, mov.get_movement_type_display(),
                                 mov.quantity, mov.reference or '-', mov.user.username if mov.user else 'Système', mov.notes or '-'])
        apply_modern_style(ws_movements)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Rapport_Global_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        return response
    finally:
        if lxml_backup: sys.modules['lxml'] = lxml_backup
        elif 'lxml' in sys.modules: del sys.modules['lxml']





@login_required
def export_reports_pdf(request):
    """Exporter les rapports en PDF"""
    import sys
    from io import BytesIO
    
    # Hack: Bypass broken lxml
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        # Date filtering
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if start_date in [None, '', 'None']:
            start_date = None
        if end_date in [None, '', 'None']:
            end_date = None
        
        # Format dates for display
        formatted_start_date = start_date if start_date else "Début"
        formatted_end_date = end_date if end_date else "Fin"
        
        # Base querysets
        invoices_qs = Invoice.objects.select_related('client').all()
        inventories_qs = Inventory.objects.select_related('product', 'product__category').all()
        
        if start_date:
            invoices_qs = invoices_qs.filter(date_issued__gte=start_date)
        
        if end_date:
            invoices_qs = invoices_qs.filter(date_issued__lte=end_date)
        
        # Préparer les données d'inventaire avec tous les calculs en Python
        inventories_data = []
        for inventory in inventories_qs[:50]:  # Limiter à 50 pour le PDF
            unit_price = float(inventory.product.selling_price)
            quantity = int(inventory.quantity)
            total_value = unit_price * quantity
            
            inventories_data.append({
                'product_name': str(inventory.product.name),
                'category_name': str(inventory.product.category.name) if inventory.product.category else '',
                'quantity': quantity,
                'unit_price': f"{unit_price:.2f}".replace('.', ','),
                'total_value': f"{total_value:.2f}".replace('.', ','),
            })
        
        # Préparer les données de factures avec tous les calculs en Python
        invoices_data = []
        status_map = {
            'paid': 'Payée',
            'sent': 'Envoyée',
            'cancelled': 'Annulée',
            'draft': 'Brouillon',
            'partial': 'Partielle'
        }
        
        for invoice in invoices_qs[:20]:  # Limiter à 20 pour le PDF
            invoice_date = invoice.date_issued.strftime('%d/%m/%Y') if invoice.date_issued else ''
            total_amount = float(invoice.total_amount)
            
            invoices_data.append({
                'invoice_date': invoice_date,
                'invoice_number': str(invoice.invoice_number),
                'client_name': str(invoice.client.name),
                'total_amount': f"{total_amount:.2f}".replace('.', ','),
                'status_display': status_map.get(invoice.status, invoice.status),
            })
        
        # Calculer les stats
        total_stock_value = float(Product.objects.aggregate(
            total=Sum(F('selling_price') * F('inventory__quantity'))
        )['total'] or Decimal('0.00'))
        
        total_products = Product.objects.count()
        
        total_sales = float(invoices_qs.filter(status='paid').aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00'))
        
        pending_sales_amount = float(invoices_qs.filter(status='sent').aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00'))
        
        # Générer le contexte
        generated_at = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        context = {
            'start_date': formatted_start_date,
            'end_date': formatted_end_date,
            'generated_at': generated_at,
            'invoices': invoices_data,
            'inventories': inventories_data,
            'total_stock_value': f"{total_stock_value:.2f}".replace('.', ','),
            'total_products': total_products,
            'total_sales': f"{total_sales:.2f}".replace('.', ','),
            'pending_sales_amount': f"{pending_sales_amount:.2f}".replace('.', ','),
        }
        
        template_path = 'inventory/reports_pdf.html'
        template = get_template(template_path)
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Rapport_Stock_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        pisa_status = pisa.CreatePDF(
            html, dest=response
        )
        
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response
    finally:
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']











# ==================== POINT OF SALE VIEWS ====================



@login_required

def pos_list(request):

    """Liste des points de vente"""

    query = request.GET.get('q', '')

    active_filter = request.GET.get('active', '')

    type_filter = request.GET.get('type', '')

    

    points_of_sale = PointOfSale.objects.select_related('manager').all()

    

    if query:

        points_of_sale = points_of_sale.filter(

            Q(name__icontains=query) | 

            Q(code__icontains=query) |

            Q(city__icontains=query)

        )

    

    if active_filter:

        is_active = active_filter == 'true'

        points_of_sale = points_of_sale.filter(is_active=is_active)

    

    if type_filter == 'warehouse':

        points_of_sale = points_of_sale.filter(is_warehouse=True)

    elif type_filter == 'store':

        points_of_sale = points_of_sale.filter(is_warehouse=False)

    

    # Annoter avec les statistiques

    points_of_sale = points_of_sale.annotate(

        product_count=Count('inventory', filter=Q(inventory__quantity__gt=0)),

        total_stock=Coalesce(Sum('inventory__quantity'), 0)

    )

    

    points_of_sale = points_of_sale.order_by('name')

    # Pagination
    paginator = Paginator(points_of_sale, 9) # 3x3 grid
    page = request.GET.get('page')
    pos_paginated = paginator.get_page(page)
    
    return render(request, 'inventory/pos/pos_list.html', {

        'points_of_sale': pos_paginated,
        'page_obj': pos_paginated,

        'query': query,

        'active_filter': active_filter,

        'type_filter': type_filter

    })





@login_required

def pos_create(request):

    """Créer un point de vente"""

    if request.method == 'POST':

        form = PointOfSaleForm(request.POST)

        if form.is_valid():

            pos = form.save()

            messages.success(request, f'Point de vente "{pos.name}" créé avec succès!')

            return redirect('inventory:pos_list')

    else:

        form = PointOfSaleForm()

    

    return render(request, 'inventory/pos/pos_form.html', {

        'form': form,

        'action': 'Créer'

    })





@login_required

def pos_update(request, pk):

    """Modifier un point de vente"""

    pos = get_object_or_404(PointOfSale, pk=pk)

    

    if request.method == 'POST':

        form = PointOfSaleForm(request.POST, instance=pos)

        if form.is_valid():

            pos = form.save()

            messages.success(request, f'Point de vente "{pos.name}" modifié avec succès!')

            return redirect('inventory:pos_detail', pk=pk)

    else:

        form = PointOfSaleForm(instance=pos)

    

    return render(request, 'inventory/pos/pos_form.html', {

        'form': form,

        'pos': pos,

        'action': 'Modifier'

    })





@login_required

def pos_detail(request, pk):

    """Détails d'un point de vente"""

    pos = get_object_or_404(PointOfSale.objects.select_related('manager'), pk=pk)

    

    # Statistiques

    inventories = pos.inventory_set.select_related('product').filter(quantity__gt=0)

    total_products = inventories.count()

    total_stock_value = pos.get_total_inventory_value()

    total_quantity = inventories.aggregate(total=Sum('quantity'))['total'] or 0

    

    # Produits en stock faible

    low_stock = inventories.filter(quantity__lte=F('reorder_level')).count()

    

    # Mouvements récents

    recent_movements = StockMovement.objects.filter(

        Q(from_point_of_sale=pos) | Q(to_point_of_sale=pos)

    ).select_related('product', 'user').order_by('-created_at')[:10]

    

    # Top produits

    top_products = inventories.order_by('-quantity')[:5]

    

    context = {

        'pos': pos,

        'total_products': total_products,

        'total_stock_value': total_stock_value,

        'total_quantity': total_quantity,

        'low_stock': low_stock,

        'recent_movements': recent_movements,

        'top_products': top_products,

    }

    

    return render(request, 'inventory/pos/pos_detail.html', context)





@login_required

def replenish_pos(request, pk):

    """Réapprovisionner un point de vente depuis l'entrepôt"""

    target_pos = get_object_or_404(PointOfSale, pk=pk)

    

    # Trouver l'entrepôt principal (Global Stock)

    warehouse = PointOfSale.objects.filter(is_warehouse=True, is_active=True).first()

    

    if not warehouse:

        messages.error(request, "Aucun entrepôt principal (Global Stock) n'est configuré.")

        return redirect('inventory:pos_detail', pk=pk)

        

    if warehouse == target_pos:

        messages.warning(request, "Impossible de réapprovisionner l'entrepôt lui-même.")

        return redirect('inventory:pos_detail', pk=pk)

    

    if request.method == 'POST':

        form = ReplenishForm(request.POST, from_pos=warehouse)

        if form.is_valid():

            # Créer manuellement le StockMovement

            product = form.cleaned_data['product']

            quantity = form.cleaned_data['quantity']

            reference = form.cleaned_data.get('reference', '')

            notes = form.cleaned_data.get('notes', '')

            

            # Si aucune référence n'est fournie, en générer une

            if not reference:

                reference = f"REPL-{datetime.now().strftime('%Y%m%d%H%M')}"

            

            try:

                movement = StockMovement.objects.create(

                    product=product,

                    movement_type='transfer',

                    quantity=quantity,

                    from_point_of_sale=warehouse,

                    to_point_of_sale=target_pos,

                    reference=reference,

                    notes=notes,

                    user=request.user

                )

                messages.success(request, f'✅ Réapprovisionnement de {target_pos.name} effectué avec succès!')

                return redirect('inventory:pos_detail', pk=pk)

            except ValidationError as e:

                messages.error(request, str(e))

    else:

        form = ReplenishForm(from_pos=warehouse)

    

    return render(request, 'inventory/movement/replenish_form.html', {

        'form': form,

        'target_pos': target_pos,

        'warehouse': warehouse

    })







@admin_required
def pos_delete(request, pk):

    """Supprimer un point de vente"""

    pos = get_object_or_404(PointOfSale, pk=pk)

    

    # Vérifier s'il y a des inventaires

    has_inventory = pos.inventory_set.filter(quantity__gt=0).exists()

    

    if request.method == 'POST':

        if has_inventory:

            messages.error(request, f'Impossible de supprimer "{pos.name}" : ce point de vente contient du stock.')

            return redirect('inventory:pos_detail', pk=pk)

        

        pos_name = pos.name

        pos.delete()

        messages.success(request, f'Point de vente "{pos_name}" supprimé avec succès!')

        return redirect('inventory:pos_list')

    

    return render(request, 'inventory/pos/pos_confirm_delete.html', {

        'pos': pos,

        'has_inventory': has_inventory

    })





# ==================== BULK STOCK CONFIGURATION ====================



@login_required

def bulk_stock_configuration(request):

    """Configuration en masse du stock pour tous les points de vente"""

    category_filter = request.GET.get('category', '')

    supplier_filter = request.GET.get('supplier', '')

    

    # Récupérer tous les points de vente actifs

    points_of_sale = PointOfSale.objects.filter(is_active=True).order_by('name')

    

    # Récupérer les produits avec filtres

    products = Product.objects.select_related('category', 'supplier').all()

    

    if category_filter:

        products = products.filter(category_id=category_filter)

    

    if supplier_filter:

        products = products.filter(supplier_id=supplier_filter)

    

    products = products.order_by('category__name', 'name')

    

    if request.method == 'POST':

        # Traiter la soumission du formulaire

        success_count = 0

        error_count = 0

        

        for product in products:

            for pos in points_of_sale:

                # Clé pour identifier chaque champ

                quantity_key = f'quantity_{product.id}_{pos.id}'

                reorder_key = f'reorder_{product.id}_{pos.id}'

                location_key = f'location_{product.id}_{pos.id}'

                

                quantity = request.POST.get(quantity_key, '').strip()

                reorder_level = request.POST.get(reorder_key, '').strip()

                location = request.POST.get(location_key, '').strip()

                

                # Si au moins un champ est rempli, créer ou mettre à jour l'inventaire

                if quantity or reorder_level or location:

                    try:

                        inventory, created = Inventory.objects.get_or_create(

                            product=product,

                            point_of_sale=pos,

                            defaults={

                                'quantity': int(quantity) if quantity else 0,

                                'reorder_level': int(reorder_level) if reorder_level else 10,

                                'location': location if location else ''

                            }

                        )

                        

                        if not created:

                            # Mettre à jour les valeurs

                            # On vérifie si la chaîne n'est pas vide pour quantity et reorder_level

                            if quantity != '':

                                inventory.quantity = int(quantity)

                            if reorder_level != '':

                                inventory.reorder_level = int(reorder_level)

                            # Pour l'emplacement, on met à jour même si c'est vide (pour permettre d'effacer)

                            # Mais on vérifie que la clé était bien dans le POST (ce qui est le cas ici)

                            inventory.location = location

                            inventory.save()

                        

                        success_count += 1

                    except Exception as e:

                        error_count += 1

        

        if success_count > 0:

            messages.success(request, f'✅ Configuration sauvegardée avec succès ! {success_count} inventaire(s) mis à jour.')

        if error_count > 0:

            messages.warning(request, f'⚠️ {error_count} erreur(s) rencontrée(s).')

        

        return redirect('inventory:bulk_stock_configuration')

    

    # Préparer les données pour le template

    # Créer un dictionnaire d'inventaires existants

    existing_inventories = {}

    for inv in Inventory.objects.filter(

        product__in=products,

        point_of_sale__in=points_of_sale

    ).select_related('product', 'point_of_sale'):

        key = f'{inv.product.id}_{inv.point_of_sale.id}'

        existing_inventories[key] = inv

    

    # Récupérer les catégories et fournisseurs pour les filtres

    categories = Category.objects.all().order_by('name')

    suppliers = Supplier.objects.all().order_by('name')

    

    context = {

        'points_of_sale': points_of_sale,

        'products': products,

        'existing_inventories': existing_inventories,

        'categories': categories,

        'suppliers': suppliers,

        'category_filter': category_filter,

        'supplier_filter': supplier_filter,

    }

    

    return render(request, 'inventory/stock/bulk_stock_config.html', context)



# ==================== API VIEWS FOR CHARTS ====================



@login_required

def api_stock_evolution(request):

    """API pour le graphique d'évolution du stock"""

    days = int(request.GET.get('days', 30))

    end_date = timezone.now()

    start_date = end_date - timedelta(days=days)

    

    # Initialiser les données

    labels = []

    entries_data = []

    exits_data = []

    

    current_date = start_date

    while current_date <= end_date:

        date_str = current_date.strftime('%d/%m')

        labels.append(date_str)

        

        # Début et fin de la journée

        day_start = current_date.replace(hour=0, minute=0, second=0, microsecond=0)

        day_end = current_date.replace(hour=23, minute=59, second=59, microsecond=999999)

        

        # Calculer les entrées

        entries = StockMovement.objects.filter(

            created_at__range=(day_start, day_end),

            movement_type__in=['entry', 'return']

        ).aggregate(total=Sum('quantity'))['total'] or 0

        entries_data.append(entries)

        

        # Calculer les sorties

        exits = StockMovement.objects.filter(

            created_at__range=(day_start, day_end),

            movement_type__in=['exit', 'defective']

        ).aggregate(total=Sum('quantity'))['total'] or 0

        exits_data.append(exits)

        

        current_date += timedelta(days=1)

    

    return JsonResponse({

        'labels': labels,

        'entries': entries_data,

        'exits': exits_data

    })



@login_required

def api_category_distribution(request):

    """API pour le graphique de répartition par catégorie"""

    categories = Category.objects.annotate(

        total_quantity=Sum('product__inventory__quantity')

    ).filter(total_quantity__gt=0).order_by('-total_quantity')

    

    labels = [cat.name for cat in categories]

    data = [cat.total_quantity for cat in categories]

    

    return JsonResponse({

        'labels': labels,

        'data': data

    })

# ==================== PRODUCT IMPORT VIEWS ====================



import sys

# Hack: Bypass broken lxml on Windows/Python 3.12 used by openpyxl

_lxml_backup = sys.modules.get('lxml')

sys.modules['lxml'] = None



try:

    import openpyxl

finally:

    # Restore lxml state

    if _lxml_backup:

        sys.modules['lxml'] = _lxml_backup

    elif 'lxml' in sys.modules:

        del sys.modules['lxml']

from django.db import transaction



@superuser_required

def product_import(request):

    """Importer des produits depuis un fichier Excel"""

    from ..forms import ProductImportForm

    

    if request.method == 'POST':

        form = ProductImportForm(request.POST, request.FILES)

        if form.is_valid():

            excel_file = request.FILES['excel_file']

            

            try:

                # Parser le fichier Excel

                workbook = openpyxl.load_workbook(excel_file)

                sheet = workbook.active

                

                # Vérifier les en-têtes

                expected_headers = ['name', 'sku', 'description', 'category', 'supplier', 'purchase_price', 'margin', 'selling_price']

                headers = [cell.value.lower().strip() if cell.value else '' for cell in sheet[1]]

                

                if not all(h in headers for h in ['name', 'sku']):

                    messages.error(request, "❌ Le fichier Excel doit contenir au minimum les colonnes 'name' et 'sku'")

                    return redirect('inventory:product_import')

                

                # Statistiques

                created_count = 0

                updated_count = 0

                error_count = 0

                errors = []

                

                # Traiter chaque ligne

                with transaction.atomic():

                    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

                        if not any(row):  # Ignorer les lignes vides

                            continue

                        

                        try:

                            # Créer un dictionnaire des données

                            row_data = dict(zip(headers, row))

                            

                            # Extraire les données

                            name = row_data.get('name', '').strip() if row_data.get('name') else None

                            sku = row_data.get('sku', '').strip() if row_data.get('sku') else None

                            

                            if not name or not sku:

                                errors.append(f"Ligne {idx}: 'name' et 'sku' sont obligatoires")

                                error_count += 1

                                continue

                            

                            # Préparer les données du produit

                            product_data = {

                                'name': name,

                                'sku': sku,

                                'description': row_data.get('description', '').strip() if row_data.get('description') else '',

                            }

                            

                            # Gérer la catégorie

                            category_name = row_data.get('category', '').strip() if row_data.get('category') else None

                            if category_name:

                                category, _ = Category.objects.get_or_create(

                                    name=category_name,

                                    defaults={'description': f'Catégorie créée automatiquement lors de l\'importation'}

                                )

                                product_data['category'] = category

                            else:

                                # Si aucune catégorie n'est fournie, utiliser une catégorie par défaut

                                default_category, _ = Category.objects.get_or_create(

                                    name='Non classé',

                                    defaults={'description': 'Catégorie par défaut pour les produits sans catégorie'}

                                )

                                product_data['category'] = default_category

                            

                            # Gérer le fournisseur

                            supplier_name = row_data.get('supplier', '').strip() if row_data.get('supplier') else None

                            if supplier_name:

                                try:

                                    supplier = Supplier.objects.get(name=supplier_name)

                                    product_data['supplier'] = supplier

                                except Supplier.DoesNotExist:

                                    errors.append(f"Ligne {idx}: Fournisseur '{supplier_name}' introuvable")

                            

                            # Gérer les prix

                            try:

                                if row_data.get('purchase_price'):

                                    product_data['purchase_price'] = Decimal(str(row_data['purchase_price']))

                                

                                if row_data.get('margin'):

                                    product_data['margin'] = Decimal(str(row_data['margin']))

                                

                                if row_data.get('selling_price'):

                                    product_data['selling_price'] = Decimal(str(row_data['selling_price']))

                                elif product_data.get('purchase_price') and product_data.get('margin'):

                                    # Calculer le prix de vente automatiquement: PV = PA + marge (montant)

                                    purchase_price = product_data['purchase_price']

                                    margin = product_data['margin']

                                    product_data['selling_price'] = purchase_price + margin

                                else:

                                    # Si aucun prix de vente n'est fourni et qu'on ne peut pas le calculer,

                                    # utiliser une valeur par défaut minimale (0.01 est le minimum requis par le modèle)

                                    product_data['selling_price'] = Decimal('0.01')

                                    errors.append(f"Ligne {idx}: Prix de vente manquant - valeur par défaut (0.01 GNF) appliquée")

                            except (ValueError, TypeError, InvalidOperation) as e:

                                errors.append(f"Ligne {idx}: Erreur de prix - {str(e)}")

                                error_count += 1

                                continue

                            

                            # Créer ou mettre à jour le produit

                            product, created = Product.objects.update_or_create(

                                sku=sku,

                                defaults=product_data

                            )

                            

                            if created:

                                created_count += 1

                                # Créer des inventaires pour tous les points de vente actifs

                                for pos in PointOfSale.objects.filter(is_active=True):

                                    Inventory.objects.create(

                                        product=product,

                                        point_of_sale=pos,

                                        quantity=0,

                                        reorder_level=10

                                    )

                            else:

                                updated_count += 1

                        

                        except Exception as e:

                            errors.append(f"Ligne {idx}: {str(e)}")

                            error_count += 1

                

                # Afficher les résultats

                if created_count > 0:

                    messages.success(request, f"✅ {created_count} produit(s) créé(s) avec succès!")

                if updated_count > 0:

                    messages.info(request, f"ℹ️ {updated_count} produit(s) mis à jour")

                if error_count > 0:

                    messages.warning(request, f"⚠️ {error_count} erreur(s) détectée(s)")

                

                # Stocker les erreurs dans la session pour affichage

                if errors:

                    request.session['import_errors'] = errors[:50]  # Limiter à 50 erreurs

                

                return redirect('inventory:product_import')

                

            except Exception as e:

                messages.error(request, f"❌ Erreur lors de la lecture du fichier: {str(e)}")

    else:

        form = ProductImportForm()

    

    # Récupérer les erreurs de la session

    import_errors = request.session.pop('import_errors', [])

    

    return render(request, 'inventory/product/product_import.html', {

        'form': form,

        'import_errors': import_errors

    })





@superuser_required

def download_product_template(request):

    """Télécharger un modèle Excel pour l'importation de produits"""

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    

    # Créer un nouveau workbook

    wb = Workbook()

    ws = wb.active

    ws.title = "Produits"

    

    # En-têtes

    headers = ['name', 'sku', 'description', 'category', 'supplier', 'purchase_price', 'margin', 'selling_price']

    header_names = ['Nom*', 'SKU*', 'Description', 'Catégorie', 'Fournisseur', 'Prix d\'achat', 'Marge (GNF)', 'Prix de vente']

    

    # Style pour les en-têtes

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF")

    header_alignment = Alignment(horizontal="center", vertical="center")

    

    # Ajouter les en-têtes

    for col_num, (header, header_name) in enumerate(zip(headers, header_names), 1):

        cell = ws.cell(row=1, column=col_num)

        cell.value = header_name

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = header_alignment

    

    # Ajouter des exemples

    examples = [

        ['Ordinateur Portable Dell', 'DELL-LAP-001', 'Ordinateur portable 15 pouces', 'Électronique', 'Dell Inc', 800.00, 25.00, 1000.00],

        ['Souris sans fil Logitech', 'LOG-MOU-001', 'Souris optique sans fil', 'Accessoires', 'Logitech', 15.00, 50.00, 22.50],

        ['Clavier mécanique RGB', 'KEY-RGB-001', 'Clavier gaming RGB', 'Accessoires', '', 45.00, 40.00, 63.00],

    ]

    

    for row_num, example in enumerate(examples, 2):

        for col_num, value in enumerate(example, 1):

            ws.cell(row=row_num, column=col_num, value=value)

    

    # Ajuster la largeur des colonnes

    column_widths = [25, 18, 35, 18, 20, 15, 12, 15]

    for col_num, width in enumerate(column_widths, 1):

        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    

    # Créer la réponse HTTP

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response['Content-Disposition'] = 'attachment; filename=modele_import_produits.xlsx'

    

    wb.save(response)

    return response





@admin_required
def reset_data(request):

    """

    Réinitialisation complète des données de l'application

    Supprime : Produits, Stocks, Mouvements, Factures, Devis, Réceptions, Paiements, Clients, Fournisseurs

    Conserve : Utilisateurs, Paramètres, Catégories, Points de Vente

    """

    if request.method == 'POST':

        confirm = request.POST.get('confirm')

        if confirm == 'RESET':

            try:

                # 1. Transactions Sales

                Payment.objects.all().delete()

                InvoiceItem.objects.all().delete()

                Invoice.objects.all().delete()

                QuoteItem.objects.all().delete()

                Quote.objects.all().delete()

                # 1b. Finance
                MonthlyProfitReport.objects.all().delete()
                Expense.objects.all().delete()

                

                # 2. Transactions Purchases

                ReceiptItem.objects.all().delete()

                Receipt.objects.all().delete()

                

                # 3. Stock

                StockMovement.objects.all().delete()

                Inventory.objects.all().delete()

                

                # 4. Master Data (Products, Clients, Suppliers)

                Product.objects.all().delete()

                Client.objects.all().delete()

                Supplier.objects.all().delete()

                

                messages.success(request, '✅ Toutes les données ont été réinitialisées avec succès. Le système est remis à zéro.')

                return redirect('inventory:settings')

            except ProtectedError as e:

                messages.error(request, f"❌ Impossible de supprimer certaines données car elles sont liées à d'autres éléments protégés : {str(e)}")

            except Exception as e:

                messages.error(request, f"❌ Une erreur est survenue lors de la réinitialisation : {str(e)}")

        else:
            messages.error(request, "❌ Confirmation incorrecte. Veuillez taper 'RESET' pour confirmer.")
            
    return render(request, 'inventory/reset_data_confirm.html')


# ==================== ADVANCED REPORTS VIEWS ====================

@staff_required
def advanced_reports_view(request):
    """Vue principale pour les rapports avancés"""
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # Récupérer les paramètres de filtrage
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    specific_date = request.GET.get('date')
    report_type = request.GET.get('report_type', 'all')
    
    if specific_date:
        start_date = specific_date
        end_date = specific_date
    
    # Convertir les dates si elles existent
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Récupérer les paramètres de l'entreprise
    company_settings = Settings.objects.first()
    if not company_settings:
        company_settings = Settings.objects.create()
    
    # 1. Activités récentes des ventes
    sales_activities = InvoiceItem.objects.select_related('invoice', 'invoice__client', 'product').all()
    if start_date:
        sales_activities = sales_activities.filter(invoice__date_issued__gte=start_date)
    if end_date:
        sales_activities = sales_activities.filter(invoice__date_issued__lte=end_date)
    sales_activities = sales_activities.order_by('-invoice__date_issued')[:50]
    
    # 2. État des factures
    invoices_paid = Invoice.objects.filter(status='paid')
    invoices_pending = Invoice.objects.filter(status__in=['draft', 'sent'])
    invoices_overdue = Invoice.objects.filter(status__in=['draft', 'sent'], date_due__lt=timezone.now().date())
    
    if start_date:
        invoices_paid = invoices_paid.filter(date_issued__gte=start_date)
        invoices_pending = invoices_pending.filter(date_issued__gte=start_date)
        invoices_overdue = invoices_overdue.filter(date_issued__gte=start_date)
    if end_date:
        invoices_paid = invoices_paid.filter(date_issued__lte=end_date)
        invoices_pending = invoices_pending.filter(date_issued__lte=end_date)
        invoices_overdue = invoices_overdue.filter(date_issued__lte=end_date)
    
    # 3. Répartition du stock par magasin
    stock_by_pos = Inventory.objects.select_related('product', 'point_of_sale').values(
        'point_of_sale__name', 'point_of_sale__code'
    ).annotate(
        total_products=Count('product', distinct=True),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('product__purchase_price'))
    ).order_by('point_of_sale__name')
    
    # 4. Produits en stock faible
    low_stock_products = Inventory.objects.select_related('product', 'point_of_sale').filter(
        quantity__lte=F('reorder_level')
    ).order_by('quantity')[:50]
    
    # 5. Derniers mouvements de stock
    stock_movements = StockMovement.objects.select_related('product', 'from_point_of_sale', 'user').all()
    if start_date:
        stock_movements = stock_movements.filter(created_at__date__gte=start_date)
    if end_date:
        stock_movements = stock_movements.filter(created_at__date__lte=end_date)
    stock_movements = stock_movements.order_by('-created_at')[:50]
    
    # 6. Produits retournés
    returned_products = StockMovement.objects.select_related('product', 'from_point_of_sale', 'user').filter(
        movement_type='return'
    )
    if start_date:
        returned_products = returned_products.filter(created_at__date__gte=start_date)
    if end_date:
        returned_products = returned_products.filter(created_at__date__lte=end_date)
    returned_products = returned_products.order_by('-created_at')[:50]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,
        'company_settings': company_settings,
        'sales_activities': sales_activities,
        'invoices_paid': invoices_paid,
        'invoices_pending': invoices_pending,
        'invoices_overdue': invoices_overdue,
        'stock_by_pos': stock_by_pos,
        'low_stock_products': low_stock_products,
        'stock_movements': stock_movements,
        'returned_products': returned_products,
    }
    
    return render(request, 'inventory/advanced_reports.html', context)


@staff_required
def export_sales_activities_excel(request):
    """Exporter les activités de vente en Excel"""
    from ..excel_utils import export_to_excel
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    specific_date = request.GET.get('date')

    if specific_date:
        start_date = specific_date
        end_date = specific_date
    
    sales_activities = InvoiceItem.objects.select_related('invoice', 'invoice__client', 'product').all()
    if start_date:
        sales_activities = sales_activities.filter(invoice__date_issued__gte=start_date)
    if end_date:
        sales_activities = sales_activities.filter(invoice__date_issued__lte=end_date)
    sales_activities = sales_activities.order_by('-invoice__date_issued')
    
    headers = ['Date', 'N° Facture', 'Client', 'Produit', 'Quantité', 'Prix Unitaire', 'Total']
    data = []
    
    for item in sales_activities:
        data.append([
            item.invoice.date_issued.strftime('%d/%m/%Y'),
            item.invoice.invoice_number,
            item.invoice.client.name,
            item.product.name,
            item.quantity,
            float(item.unit_price),
            float(item.total)
        ])
    
    return export_to_excel(headers, data, "Activités de Vente", "Activites_Ventes")


@staff_required
def export_sales_activities_pdf(request):
    """Exporter les activités de vente en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        sales_activities = InvoiceItem.objects.select_related('invoice', 'invoice__client', 'product').all()
        if start_date:
            sales_activities = sales_activities.filter(invoice__date_issued__gte=start_date)
        if end_date:
            sales_activities = sales_activities.filter(invoice__date_issued__lte=end_date)
        sales_activities = sales_activities.order_by('-invoice__date_issued')[:100]
        
        company_settings = Settings.objects.first()
        context = {
            'report_title': 'Activités Récentes des Ventes',
            'sales_activities': sales_activities,
            'start_date': start_date,
            'end_date': end_date,
            'company_settings': company_settings,
            'generated_at': timezone.now(),
        }
        
        template = get_template('inventory/reports_pdf/sales_activities_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Activites_Ventes_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']


@superuser_required
def export_invoice_status_excel(request):
    """Exporter l'état des factures en Excel"""
    from ..excel_utils import export_to_excel
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    specific_date = request.GET.get('date')

    if specific_date:
        start_date = specific_date
        end_date = specific_date
    
    query = request.GET.get('q', '') or request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    invoices = Invoice.objects.select_related('client').all()
    
    # Filtrage par point de vente (Sécurité)
    invoices = filter_queryset_by_pos(invoices, request.user, 'point_of_sale')
    
    if query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=query) | 
            Q(client__name__icontains=query)
        )
    
    if status_filter:
        invoices = invoices.filter(status=status_filter)

    if start_date:
        invoices = invoices.filter(date_issued__gte=start_date)
    if end_date:
        invoices = invoices.filter(date_issued__lte=end_date)
    
    headers = ['N° Facture', 'Client', 'Date Émission', 'Date Échéance', 'Statut', 'Montant Total', 'Montant Payé', 'Solde']
    data = []
    
    for invoice in invoices:
        data.append([
            invoice.invoice_number,
            invoice.client.name,
            invoice.date_issued.strftime('%d/%m/%Y'),
            invoice.date_due.strftime('%d/%m/%Y'),
            invoice.get_status_display(),
            float(invoice.total_amount),
            float(invoice.get_amount_paid()),
            float(invoice.get_remaining_amount())
        ])
    
    # Calculate totals
    total_amount_sum = sum(invoice.total_amount for invoice in invoices)
    total_paid_sum = sum(invoice.get_amount_paid() for invoice in invoices)
    total_remaining_sum = sum(invoice.get_remaining_amount() for invoice in invoices)
    
    # Add totals row
    data.append(['', '', '', '', 'TOTAL GÉNÉRAL', float(total_amount_sum), float(total_paid_sum), float(total_remaining_sum)])
    
    return export_to_excel(headers, data, "État des Factures", "Etat_Factures")


@staff_required
@superuser_required
def export_invoice_status_pdf(request):
    """Exporter l'état des factures en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        specific_date = request.GET.get('date')

        if specific_date:
            start_date = specific_date
            end_date = specific_date
        specific_date = request.GET.get('date')
        query = request.GET.get('q', '') or request.GET.get('search', '')
        status_filter = request.GET.get('status', '')

        if specific_date:
            start_date = specific_date
            end_date = specific_date
        
        invoices_qs = Invoice.objects.select_related('client').all()
        
        # Filtrage par point de vente (Sécurité)
        invoices_qs = filter_queryset_by_pos(invoices_qs, request.user, 'point_of_sale')
        
        if query:
            invoices_qs = invoices_qs.filter(
                Q(invoice_number__icontains=query) | 
                Q(client__name__icontains=query)
            )
        
        if status_filter:
            invoices_qs = invoices_qs.filter(status=status_filter)

        if start_date:
            invoices_qs = invoices_qs.filter(date_issued__gte=start_date)
        if end_date:
            invoices_qs = invoices_qs.filter(date_issued__lte=end_date)
        invoices_qs = invoices_qs.order_by('-date_issued')
        
        # Pré-calculer toutes les données pour le PDF
        status_map = {
            'paid': 'Payée',
            'sent': 'Envoyée',
            'cancelled': 'Annulée',
            'draft': 'Brouillon',
            'partial': 'Partielle'
        }
        
        invoices_data = []
        total_amount_sum = 0
        total_remaining_sum = 0
        
        for invoice in invoices_qs:
            total_amount = float(invoice.total_amount)
            remaining = float(invoice.get_remaining_amount())
            
            total_amount_sum += total_amount
            total_remaining_sum += remaining
            
            invoices_data.append({
                'invoice_number': invoice.invoice_number,
                'client_name': invoice.client.name,
                'date_issued': invoice.date_issued,
                'date_due': invoice.date_due,
                'status_display': status_map.get(invoice.status, invoice.status),
                'status': invoice.status,
                'total_amount': total_amount,
                'remaining_amount': remaining,
            })
        
        company_settings = Settings.objects.first()
        
        # Formater les dates
        formatted_start_date = start_date if start_date else None
        formatted_end_date = end_date if end_date else None
        generated_at = datetime.now().strftime('%d/%m/%Y à %H:%M')
        
        context = {
            'report_title': 'État des Factures',
            'invoices': invoices_data,
            'start_date': start_date,
            'end_date': end_date,
            'company_settings': company_settings,
            'generated_at': datetime.now(),
            'total_amount_sum': total_amount_sum,
            'total_remaining_sum': total_remaining_sum,
        }
        
        template = get_template('inventory/reports_pdf/invoice_status_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Etat_Factures_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']


@superuser_required
def export_stock_distribution_excel(request):
    """Exporter la répartition du stock par magasin en Excel"""
    from ..excel_utils import export_to_excel
    
    stock_by_pos = Inventory.objects.select_related('product', 'point_of_sale').values(
        'point_of_sale__name', 'point_of_sale__code'
    ).annotate(
        total_products=Count('product', distinct=True),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('product__purchase_price'))
    ).order_by('point_of_sale__name')
    
    headers = ['Point de Vente', 'Code', 'Nombre de Produits', 'Quantité Totale', 'Valeur Totale']
    data = []
    
    for item in stock_by_pos:
        data.append([
            item['point_of_sale__name'],
            item['point_of_sale__code'],
            item['total_products'],
            item['total_quantity'] or 0,
            float(item['total_value'] or 0)
        ])
    
    return export_to_excel(headers, data, "Répartition Stock", "Repartition_Stock")


@superuser_required
def export_stock_distribution_pdf(request):
    """Exporter la répartition du stock par magasin en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        stock_by_pos = Inventory.objects.select_related('product', 'point_of_sale').values(
            'point_of_sale__name', 'point_of_sale__code'
        ).annotate(
            total_products=Count('product', distinct=True),
            total_quantity=Sum('quantity'),
            total_value=Sum(F('quantity') * F('product__purchase_price'))
        ).order_by('point_of_sale__name')
        
        company_settings = Settings.objects.first()
        # Pré-calculer/formatter les valeurs pour un rendu PDF fiable
        stock_by_pos_list = []
        for item in stock_by_pos:
            total_products = item.get('total_products') or 0
            total_quantity = item.get('total_quantity') or 0
            total_value = float(item.get('total_value') or 0)
            stock_by_pos_list.append({
                'point_of_sale_name': item.get('point_of_sale__name') or '',
                'point_of_sale_code': item.get('point_of_sale__code') or '',
                'total_products': int(total_products),
                'total_quantity': int(total_quantity),
                'total_value': f"{total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.' ) + f" {getattr(company_settings, 'currency', 'GNF')}",
            })

        context = {
            'report_title': 'Répartition du Stock par Magasin',
            'stock_by_pos': stock_by_pos_list,
            'company_settings': company_settings,
            'generated_at': timezone.now(),
        }
        
        template = get_template('inventory/reports_pdf/stock_distribution_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Repartition_Stock_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']


@staff_required
def export_low_stock_excel(request):
    """Exporter les produits en stock faible en Excel"""
    from ..excel_utils import export_to_excel
    
    low_stock = Inventory.objects.select_related('product', 'point_of_sale').filter(
        quantity__lte=F('reorder_level')
    ).order_by('quantity')
    
    headers = ['Produit', 'SKU', 'Point de Vente', 'Quantité Actuelle', 'Seuil de Réapprovisionnement', 'Emplacement']
    data = []
    
    for item in low_stock:
        data.append([
            item.product.name,
            item.product.sku or '-',
            item.point_of_sale.name,
            item.quantity,
            item.reorder_level,
            item.location or '-'
        ])
    
    return export_to_excel(headers, data, "Stock Faible", "Stock_Faible")


@staff_required
def export_low_stock_pdf(request):
    """Exporter les produits en stock faible en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        low_stock = Inventory.objects.select_related('product', 'point_of_sale').filter(
            quantity__lte=F('reorder_level')
        ).order_by('quantity')[:100]
        
        company_settings = Settings.objects.first()
        context = {
            'report_title': 'Produits en Stock Faible',
            'low_stock_products': low_stock,
            'company_settings': company_settings,
            'generated_at': timezone.now(),
        }
        
        template = get_template('inventory/reports_pdf/low_stock_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Stock_Faible_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']


@staff_required
def export_stock_movements_excel(request):
    """Exporter les mouvements de stock en Excel"""
    from ..excel_utils import export_to_excel
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search', '')
    movement_type = request.GET.get('type', '')
    specific_date = request.GET.get('date', '')
    user_id = request.GET.get('user')
    product_id = request.GET.get('product')

    if specific_date:
        start_date = specific_date
        end_date = specific_date
    
    movements = StockMovement.objects.select_related('product', 'from_point_of_sale', 'user').all()
    
    # Filtre par point de vente pour STAFF
    movements = filter_queryset_by_pos(movements, request.user, 'from_point_of_sale')

    if search_query:
        movements = movements.filter(
            Q(product__name__icontains=search_query) | 
            Q(product__sku__icontains=search_query) |
            Q(reference__icontains=search_query) |
            Q(notes__icontains=search_query)
        ).distinct()

    if movement_type:
        movements = movements.filter(movement_type=movement_type)

    if user_id:
        movements = movements.filter(user_id=user_id)

    if product_id:
        movements = movements.filter(product_id=product_id)

    if start_date:
        movements = movements.filter(created_at__date__gte=start_date)
    if end_date:
        movements = movements.filter(created_at__date__lte=end_date)
    movements = movements.order_by('-created_at')
    
    headers = ['Date', 'Produit', 'Type', 'Quantité', 'Point de Vente', 'Référence', 'Utilisateur', 'Notes']
    data = []
    
    for movement in movements:
        data.append([
            movement.created_at.strftime('%d/%m/%Y %H:%M'),
            movement.product.name,
            movement.get_movement_type_display(),
            movement.quantity,
            movement.from_point_of_sale.name,
            movement.reference or '-',
            movement.user.username if movement.user else '-',
            movement.notes or '-'
        ])
    
    return export_to_excel(headers, data, "Mouvements de Stock", "Mouvements_Stock")


@staff_required
def export_stock_movements_pdf(request):
    """Exporter les mouvements de stock en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from xhtml2pdf import pisa
        from django.template.loader import get_template

        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        search_query = request.GET.get('search', '')
        movement_type = request.GET.get('type', '')
        user_id = request.GET.get('user')
        product_id = request.GET.get('product')

        movements = StockMovement.objects.select_related('product', 'from_point_of_sale', 'user').all()
        movements = filter_queryset_by_pos(movements, request.user, 'from_point_of_sale')

        if search_query:
            movements = movements.filter(
                Q(product__name__icontains=search_query) | 
                Q(product__sku__icontains=search_query) |
                Q(reference__icontains=search_query) |
                Q(notes__icontains=search_query)
            ).distinct()

        if movement_type: movements = movements.filter(movement_type=movement_type)
        if user_id: movements = movements.filter(user_id=user_id)
        if product_id: movements = movements.filter(product_id=product_id)
        
        if start_date:
            try: movements = movements.filter(created_at__gte=datetime.strptime(start_date, '%Y-%m-%d'))
            except: pass
        if end_date:
            try: movements = movements.filter(created_at__lte=datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
            except: pass
        
        movements = movements.order_by('-created_at')[:1000]
        
        context = {
            'movements': movements,
            'company_settings': Settings.objects.first(),
            'report_title': "Mouvements de Stock",
            'generated_at': timezone.now()
        }
        
        template = get_template('inventory/reports_pdf/movement_list_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Mouvements_Stock_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err: return HttpResponse(f"Error generating PDF: {pisa_status.err}")
        return response
    finally:
        if lxml_backup: sys.modules['lxml'] = lxml_backup
        elif 'lxml' in sys.modules: del sys.modules['lxml']


@staff_required
def export_returned_products_excel(request):
    """Exporter les produits retournés en Excel"""
    from ..excel_utils import export_to_excel
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    specific_date = request.GET.get('date')

    if specific_date:
        start_date = specific_date
        end_date = specific_date
    
    returned = StockMovement.objects.select_related('product', 'from_point_of_sale', 'user').filter(
        movement_type='return'
    )
    if start_date:
        returned = returned.filter(created_at__date__gte=start_date)
    if end_date:
        returned = returned.filter(created_at__date__lte=end_date)
    returned = returned.order_by('-created_at')
    
    headers = ['Date', 'Produit', 'Quantité', 'Point de Vente', 'Référence', 'Utilisateur', 'Notes']
    data = []
    
    for item in returned:
        data.append([
            item.created_at.strftime('%d/%m/%Y %H:%M'),
            item.product.name,
            item.quantity,
            item.from_point_of_sale.name,
            item.reference or '-',
            item.user.username if item.user else '-',
            item.notes or '-'
        ])
    
    return export_to_excel(headers, data, "Produits Retournés", "Produits_Retournes")


@staff_required
def export_returned_products_pdf(request):
    """Exporter les produits retournés en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        specific_date = request.GET.get('date')

        if specific_date:
            start_date = specific_date
            end_date = specific_date
        
        returned = StockMovement.objects.select_related('product', 'from_point_of_sale', 'user').filter(
            movement_type='return'
        )
        if start_date:
            returned = returned.filter(created_at__date__gte=start_date)
        if end_date:
            returned = returned.filter(created_at__date__lte=end_date)
        returned = returned.order_by('-created_at')[:100]
        
        company_settings = Settings.objects.first()
        context = {
            'report_title': 'Produits Retournés',
            'returned_products': returned,
            'start_date': start_date,
            'end_date': end_date,
            'company_settings': company_settings,
            'generated_at': timezone.now(),
        }
        
        template = get_template('inventory/reports_pdf/returned_products_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Produits_Retournes_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup:
            sys.modules['lxml'] = lxml_backup
        else:
            if 'lxml' in sys.modules:
                del sys.modules['lxml']



@staff_required
def export_products_excel(request):
    """Exporter tous les produits en Excel"""
    from ..excel_utils import export_to_excel
    
    # Récupérer tous les produits avec leur stock total
    products = Product.objects.annotate(
        total_stock=Sum('inventory__quantity')
    ).select_related('category').order_by('name')
    
    headers = ['Nom', 'SKU', 'Catégorie', 'Stock Total', 'Colis', 'Unités', 'Analyse', 'Prix Achat', 'Prix Vente', 'Marge', 'Description']
    data = []
    
    for product in products:
        total_stock = product.total_stock or 0
        margin = float(product.selling_price - product.purchase_price) if product.selling_price and product.purchase_price else 0
        analysis_data = product.get_analysis_data()
        
        data.append([
            product.name,
            product.sku or '-',
            product.category.name if product.category else '-',
            total_stock,
            analysis_data['colis'],
            analysis_data['unites'],
            analysis_data['analysis'],
            float(product.purchase_price) if product.purchase_price else 0,
            float(product.selling_price) if product.selling_price else 0,
            margin,
            product.description or '-'
        ])
    
    return export_to_excel(headers, data, "Liste des Produits", "Liste_Produits")


@staff_required
def export_products_pdf(request):
    """Exporter tous les produits en PDF"""
    import sys
    lxml_backup = sys.modules.get('lxml')
    sys.modules['lxml'] = None
    
    try:
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        # Récupérer tous les produits avec leur stock total (limité à 200 pour le PDF)
        products = Product.objects.annotate(
            total_stock=Sum('inventory__quantity')
        ).select_related('category').order_by('name')[:200]
        
        context = {
            'report_title': 'Liste des Produits',
            'products': products,
            'company_settings': Settings.objects.first(),
            'generated_at': timezone.now(),
        }
        
        template = get_template('inventory/reports_pdf/products_list_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Liste_Produits_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    finally:
        if lxml_backup: sys.modules['lxml'] = lxml_backup
        elif 'lxml' in sys.modules: del sys.modules['lxml']
